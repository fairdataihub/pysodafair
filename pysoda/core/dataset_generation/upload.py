
from pysoda.utils.exceptions import PennsieveAgentRPCError
from ...utils import (
    generate_options_set, generating_locally, generating_on_ps, 
    uploading_with_ps_account, uploading_to_existing_ps_dataset, 
    can_resume_prior_upload, virtual_dataset_empty, PropertyNotSetError, 
    connect_pennsieve_client, get_dataset_id, get_access_token,
    PennsieveActionNoPermission, PennsieveDatasetCannotBeFound,
    EmptyDatasetError, LocalDatasetMissingSpecifiedFiles,
    PennsieveUploadException, create_request_headers, check_forbidden_characters_ps, get_users_dataset_list,
    PennsieveDatasetNameInvalid, PennsieveDatasetNameTaken, PennsieveAccountInvalid, TZLOCAL, GenerateOptionsNotSet,
    PennsieveDatasetFilesInvalid
)
from ..permissions import pennsieve_get_current_user_permissions
from os.path import isdir, isfile, getsize
from ..metadata import create_high_level_manifest_files, get_auto_generated_manifest_files, manifest, subjects, samples, code_description, dataset_description, performances, resources, sites, submission, text_metadata, METADATA_UPLOAD_PS_PATH, create_high_lvl_manifest_files_existing_ps_starting_point
from ..upload_manifests import get_upload_manifests
from .. import logger

main_curate_progress_message = ""
main_curate_status = ""

# -*- coding: utf-8 -*-

### Import required python modules
import platform
import os
from os import listdir, makedirs, mkdir, walk, rename
from os.path import (
    isdir,
    isfile,
    join,
    splitext,
    basename,
    exists,
    expanduser,
    dirname,
    getsize,
    abspath,
)
import pandas as pd
import time
from timeit import default_timer as timer
from datetime import timedelta, timezone
import shutil
import subprocess
import gevent
import pathlib
import requests
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
# from utils import connect_pennsieve_client, get_dataset_id, create_request_headers, TZLOCAL, get_users_dataset_list
# from manifest import create_high_lvl_manifest_files_existing_ps_starting_point, create_high_level_manifest_files, get_auto_generated_manifest_files
# from errors import PennsieveUploadException
from .manifestSession import UploadManifestSession
from ...constants import PENNSIEVE_URL
from ..dataset_importing import import_pennsieve_dataset

# from pysodaUtils import (
#     check_forbidden_characters_ps
# )

# from organizeDatasets import import_pennsieve_dataset


### Global variables
curateprogress = " "
curatestatus = " "
curateprintstatus = " "
total_dataset_size = 1
curated_dataset_size = 0
start_time = 0
uploaded_folder_counter = 0
current_size_of_uploaded_files = 0
generated_dataset_id = None
# the pennsieve python client used for uploading dataset files 
client = None 

userpath = expanduser("~")
configpath = join(userpath, ".pennsieve", "config.ini")
submitdataprogress = " "
submitdatastatus = " "
submitprintstatus = " "
total_file_size = 1
uploaded_file_size = 0
start_time_bf_upload = 0
start_submit = 0
metadatapath = join(userpath, "SODA", "SODA_metadata")
ps_recognized_file_extensions = [
    ".cram",
    ".jp2",
    ".jpx",
    ".lsm",
    ".ndpi",
    ".nifti",
    ".oib",
    ".oif",
    ".roi",
    ".rtf",
    ".swc",
    ".abf",
    ".acq",
    ".adicht",
    ".adidat",
    ".aedt",
    ".afni",
    ".ai",
    ".avi",
    ".bam",
    ".bash",
    ".bcl",
    ".bcl.gz",
    ".bin",
    ".brik",
    ".brukertiff.gz",
    ".continuous",
    ".cpp",
    ".csv",
    ".curv",
    ".cxls",
    ".czi",
    ".data",
    ".dcm",
    ".df",
    ".dicom",
    ".doc",
    ".docx",
    ".e",
    ".edf",
    ".eps",
    ".events",
    ".fasta",
    ".fastq",
    ".fcs",
    ".feather",
    ".fig",
    ".gif",
    ".h4",
    ".h5",
    ".hdf4",
    ".hdf5",
    ".hdr",
    ".he2",
    ".he5",
    ".head",
    ".hoc",
    ".htm",
    ".html",
    ".ibw",
    ".img",
    ".ims",
    ".ipynb",
    ".jpeg",
    ".jpg",
    ".js",
    ".json",
    ".lay",
    ".lh",
    ".lif",
    ".m",
    ".mat",
    ".md",
    ".mef",
    ".mefd.gz",
    ".mex",
    ".mgf",
    ".mgh",
    ".mgh.gz",
    ".mgz",
    ".mnc",
    ".moberg.gz",
    ".mod",
    ".mov",
    ".mp4",
    ".mph",
    ".mpj",
    ".mtw",
    ".ncs",
    ".nd2",
    ".nev",
    ".nex",
    ".nex5",
    ".nf3",
    ".nii",
    ".nii.gz",
    ".ns1",
    ".ns2",
    ".ns3",
    ".ns4",
    ".ns5",
    ".ns6",
    ".nwb",
    ".ogg",
    ".ogv",
    ".ome.btf",
    ".ome.tif",
    ".ome.tif2",
    ".ome.tif8",
    ".ome.tiff",
    ".ome.xml",
    ".openephys",
    ".pdf",
    ".pgf",
    ".png",
    ".ppt",
    ".pptx",
    ".ps",
    ".pul",
    ".py",
    ".r",
    ".raw",
    ".rdata",
    ".rh",
    ".rhd",
    ".sh",
    ".sldasm",
    ".slddrw",
    ".smr",
    ".spikes",
    ".svg",
    ".svs",
    ".tab",
    ".tar",
    ".tar.gz",
    ".tcsh",
    ".tdm",
    ".tdms",
    ".text",
    ".tif",
    ".tiff",
    ".tsv",
    ".txt",
    ".vcf",
    ".webm",
    ".xlsx",
    ".xml",
    ".yaml",
    ".yml",
    ".zip",
    ".zsh",
]

myds = ""
initial_bfdataset_size = 0
upload_directly_to_bf = 0
initial_bfdataset_size_submit = 0
renaming_files_flow = False

total_files = 0 # the total number of files in a given dataset that need to be uploaded to Pennsieve
total_bytes_uploaded = 0 # current number of bytes uploaded to Pennsieve in the upload session
total_upload_size = 0 # total number of bytes to upload to Pennsieve in the upload session

forbidden_characters = '<>:"/\|?*'
forbidden_characters_bf = '\/:*?"<>'

# a global that tracks the amount of files that have been uploaded in an upload session;
# is reset once the session ends by success, or failure (is implicitly reset in case of Pennsieve Agent freeze by the user closing SODA)
main_curation_uploaded_files = 0

DEV_TEMPLATE_PATH = join(dirname(__file__), "..", "file_templates")

# once pysoda has been packaged with pyinstaller
# it becomes nested into the pysodadist/api directory
PROD_TEMPLATE_PATH = join(dirname(__file__), "..", "..", "file_templates")
TEMPLATE_PATH = DEV_TEMPLATE_PATH if exists(DEV_TEMPLATE_PATH) else PROD_TEMPLATE_PATH




ums = UploadManifestSession()





def open_file(file_path):
    """
    Opening folder on all platforms
    https://stackoverflow.com/questions/6631299/python-opening-a-folder-in-explorer-nautilus-mac-thingie

    Args:
        file_path: path of the folder (string)
    Action:
        Opens file explorer window to the given path
    """

    if platform.system() == "Windows":
        subprocess.Popen(f"explorer /select,{str(file_path)}")
    elif platform.system() == "Darwin":
        subprocess.Popen(["open", file_path])
    else:
        subprocess.Popen(["xdg-open", file_path])



def folder_size(path):
    """
    Provides the size of the folder indicated by path

    Args:
        path: path of the folder (string)
    Returns:
        total_size: total size of the folder in bytes (integer)
    """
    total_size = 0

    for path, dirs, files in walk(path):
        for f in files:
            fp = join(path, f)
            total_size += getsize(fp)
    return total_size


def path_size(path):
    """
    Returns size of the path, after checking if it's a folder or a file
    Args:
        path: path of the file/folder (string)
    Returns:
        total_size: total size of the file/folder in bytes (integer)
    """
    return folder_size(path) if isdir(path) else getsize(path)



def return_new_path(topath):
    """
    This function checks if a folder already exists and in such cases,
    appends (1) or (2) etc. to the folder name

    Args:
        topath: path where the folder is supposed to be created (string)
    Returns:
        topath: new folder name based on the availability in destination folder (string)
    """

    if not exists(topath):
        return topath

    i = 1
    while True:
        if not exists(topath + " (" + str(i) + ")"):
            return topath + " (" + str(i) + ")"
        i += 1


def return_new_path_replace(topath):
    """
    This function checks if a folder already exists and in such cases,
    replace the existing folder (this is the opposite situation to the function return_new_path)

    Args:
        topath: path where the folder is supposed to be created (string)
    Returns:
        topath: new folder name based on the availability in destination folder (string)
    """

    if not exists(topath):
        return topath
    i = 1
    while True:
        if not exists(topath + " (" + str(i) + ")"):
            return topath + " (" + str(i) + ")"
        i += 1


def time_format(elapsed_time):
    mins, secs = divmod(elapsed_time, 60)
    hours, mins = divmod(mins, 60)
    return "%dh:%02dmin:%02ds" % (hours, mins, secs)


def mycopyfileobj(fsrc, fdst, length=16 * 1024 * 16):
    """
    Helper function to copy file

    Args:
        fsrc: source file opened in python (file-like object)
        fdst: destination file accessed in python (file-like object)
        length: copied buffer size in bytes (integer)
    """
    global curateprogress
    global total_dataset_size
    global curated_dataset_size
    global main_generated_dataset_size

    while True:
        buf = fsrc.read(length)
        if not buf:
            break
        gevent.sleep(0)
        fdst.write(buf)
        curated_dataset_size += len(buf)
        main_generated_dataset_size += len(buf)


def mycopyfile_with_metadata(src, dst, *, follow_symlinks=True):
    """
    Copy file src to dst with metadata (timestamp, permission, etc.) conserved

    Args:
        src: source file (string)
        dst: destination file (string)
    Returns:
        dst
    """
    if not follow_symlinks and os.path.islink(src):
        os.symlink(os.readlink(src), dst)
    else:
        with open(src, "rb") as fsrc:
            with open(dst, "wb") as fdst:
                mycopyfileobj(fsrc, fdst)
    shutil.copystat(src, dst)
    return dst


def check_empty_files_folders(soda):
    """
    Function to check for empty files and folders

    Args:
        soda: soda dict with information about all specified files and folders
    Output:
        error: error message with list of non valid local data files, if any
    """
    try:
        def recursive_empty_files_check(my_folder, my_relative_path, error_files):
            for folder_key, folder in my_folder["folders"].items():
                relative_path = my_relative_path + "/" + folder_key
                error_files = recursive_empty_files_check(
                    folder, relative_path, error_files
                )

            for file_key in list(my_folder["files"].keys()):
                file = my_folder["files"][file_key]
                file_type = file.get("location")
                if file_type == "local":
                    file_path = file["path"]
                    if isfile(file_path):
                        file_size = getsize(file_path)
                        if file_size == 0:
                            del my_folder["files"][file_key]
                            relative_path = my_relative_path + "/" + file_key
                            error_message = relative_path + " (path: " + file_path + ")"
                            error_files.append(error_message)

            return error_files

        def recursive_empty_local_folders_check(
                    my_folder,
                    my_folder_key,
                    my_folders_content,
                    my_relative_path,
                    error_folders,
                ):
            folders_content = my_folder["folders"]
            for folder_key in list(my_folder["folders"].keys()):
                folder = my_folder["folders"][folder_key]
                relative_path = my_relative_path + "/" + folder_key
                error_folders = recursive_empty_local_folders_check(
                    folder, folder_key, folders_content, relative_path, error_folders
                )

            if not my_folder["folders"] and not my_folder["files"]:
                ignore = False
                if "location" in my_folder and my_folder.get("location") == "ps":
                    ignore = True
                if not ignore:
                    error_message = my_relative_path
                    error_folders.append(error_message)
                    del my_folders_content[my_folder_key]
            return error_folders

        error_files = []
        error_folders = []
        if "dataset-structure" in soda.keys():
            dataset_structure = soda["dataset-structure"]
            if "folders" in dataset_structure:
                for folder_key, folder in dataset_structure["folders"].items():
                    relative_path = folder_key
                    error_files = recursive_empty_files_check(
                        folder, relative_path, error_files
                    )

                folders_content = dataset_structure["folders"]
                for folder_key in list(dataset_structure["folders"].keys()):
                    folder = dataset_structure["folders"][folder_key]
                    relative_path = folder_key
                    error_folders = recursive_empty_local_folders_check(
                        folder,
                        folder_key,
                        folders_content,
                        relative_path,
                        error_folders,
                    )

        if len(error_files) > 0:
            error_message = [
                "The following local file(s) is/are empty (0 kb) and will be ignored."
            ]
            error_files = error_message + [] + error_files

        if len(error_folders) > 0:
            error_message = [
                "The SPARC dataset structure does not allow empty folders. The following empty folders will be removed from your dataset:"
            ]
            error_folders = error_message + [] + error_folders

        return {
            "empty_files": error_files, 
            "empty_folders": error_folders, 
            "soda": soda
        }

    except Exception as e:
        raise e


def check_local_dataset_files_validity(soda):
    """
    Function to check that the local data files and folders specified in the dataset are valid

    Args:
        soda: soda dict with information about all specified files and folders
    Output:
        error: error message with list of non valid local data files, if any
    """

    def recursive_local_file_check(my_folder, my_relative_path, error):
        for folder_key, folder in my_folder["folders"].items():
            relative_path = my_relative_path + "/" + folder_key
            error = recursive_local_file_check(folder, relative_path, error)

        for file_key in list(my_folder["files"].keys()):
            file = my_folder["files"][file_key]
            if file_key in ["manifest.xlsx", "manifest.csv"]:
                continue
            file_type = file.get("location")
            if file_type == "local":
                file_path = file["path"]
                if file.get("location") == "ps":
                    continue
                if not isfile(file_path):
                    relative_path = my_relative_path + "/" + file_key
                    error_message = relative_path + " (path: " + file_path + ")"
                    error.append(error_message)
                else:
                    file_size = getsize(file_path)
                    if file_size == 0:
                        del my_folder["files"][file_key]

        return error

    def recursive_empty_local_folder_remove(
        my_folder, my_folder_key, my_folders_content
    ):

        folders_content = my_folder["folders"]
        for folder_key in list(my_folder["folders"].keys()):
            folder = my_folder["folders"][folder_key]
            recursive_empty_local_folder_remove(folder, folder_key, folders_content)

        if not my_folder.get("folders") and not my_folder.get("files") and my_folder.get("location") != "ps":
            del my_folders_content[my_folder_key]

    error = []
    if "dataset-structure" in soda.keys():
        dataset_structure = soda["dataset-structure"]
        # Remove 0kb files, files that can't be found, and any empty folders from the dataset data files
        if "folders" in dataset_structure:
            for folder_key, folder in dataset_structure["folders"].items():
                relative_path = folder_key
                error = recursive_local_file_check(folder, relative_path, error)

            folders_content = dataset_structure["folders"]
            for folder_key in list(dataset_structure["folders"].keys()):
                folder = dataset_structure["folders"][folder_key]
                recursive_empty_local_folder_remove(folder, folder_key, folders_content)

    # Return list of all the files that were not found. 
    if len(error) > 0:
        error_message = [
            "Error: The following local files were not found. Specify them again or remove them."
        ]
        error = error_message + error

    return error


# path to local SODA folder for saving manifest files
manifest_sparc = ["manifest.xlsx", "manifest.csv"]
manifest_folder_path = join(userpath, ".pysoda", "manifest_file")



def check_json_size(jsonStructure):
    """
        This function is called to check size of files that will be created locally on a user's device.
    """
    global total_dataset_size
    total_dataset_size = 0

    try:
        def recursive_dataset_scan(folder):
            global total_dataset_size

            if "files" in folder.keys():
                for file_key, file in folder["files"].items():
                    if "deleted" not in file["action"]:
                        file_type = file.get("location")
                        if file_type == "local":
                            file_path = file["path"]
                            if isfile(file_path):
                                total_dataset_size += getsize(file_path)

            if "folders" in folder.keys():
                for folder_key, folder in folder["folders"].items():
                    recursive_dataset_scan(folder)

        # scan dataset structure
        dataset_structure = jsonStructure["dataset-structure"]
        folderSection = dataset_structure["folders"]
        # gets keys like code, primary, source and their content...
        for keys, contents in folderSection.items():
            recursive_dataset_scan(contents)

        if "manifest-files" in jsonStructure.keys():
            manifest_files_structure = create_high_level_manifest_files(jsonStructure, manifest_folder_path)
            for key in manifest_files_structure.keys():
                manifestpath = manifest_files_structure[key]
                if isfile(manifestpath):
                    total_dataset_size += getsize(manifestpath)

        # returns in bytes
        return {"dataset_size": total_dataset_size}
    except Exception as e:
        raise e


def generate_dataset_locally(soda):
    global logger
    logger.info("starting generate_dataset_locally")

    # Vars used for tracking progress on the frontend 
    global main_curate_progress_message
    global progress_percentage
    global main_total_generate_dataset_size
    global start_generate
    global main_curation_uploaded_files

    main_curation_uploaded_files = 0

    def recursive_dataset_scan(
        my_folder, my_folderpath, list_copy_files, list_move_files
    ):
        global main_total_generate_dataset_size

        if "folders" in my_folder.keys():
            for folder_key, folder in my_folder["folders"].items():
                folderpath = join(my_folderpath, folder_key)
                if not isdir(folderpath):
                    mkdir(folderpath)
                list_copy_files, list_move_files = recursive_dataset_scan(
                    folder, folderpath, list_copy_files, list_move_files
                )

        if "files" in my_folder.keys():
            for file_key, file in my_folder["files"].items():
                if "deleted" not in file["action"]:
                    file_type = file.get("location")
                    if file_type == "local":
                        file_path = file["path"]
                        if isfile(file_path):
                            destination_path = abspath(
                                join(my_folderpath, file_key)
                            )
                            if not isfile(destination_path):
                                if (
                                    "existing" in file["action"]
                                    and soda["generate-dataset"][
                                        "if-existing"
                                    ]
                                    == "merge"
                                ):
                                    list_move_files.append(
                                        [file_path, destination_path]
                                    )
                                else:
                                    main_total_generate_dataset_size += getsize(
                                        file_path
                                    )
                                    list_copy_files.append(
                                        [file_path, destination_path]
                                    )
                        else:
                            logger.info(f"file_path {file_path} does not exist. Skipping.")
        return list_copy_files, list_move_files


    logger.info("generate_dataset_locally step 1")
    # 1. Create new folder for dataset or use existing merge with existing or create new dataset
    main_curate_progress_message = "Generating folder structure and list of files to be included in the dataset"
    dataset_absolute_path = soda["generate-dataset"]["path"]
    if_existing = soda["generate-dataset"]["if-existing"]
    dataset_name = soda["generate-dataset"]["dataset-name"]
    datasetpath = join(dataset_absolute_path, dataset_name)
    datasetpath = return_new_path(datasetpath)
    mkdir(datasetpath)

    logger.info("generate_dataset_locally step 2")
    # 2. Scan the dataset structure and:
    # 2.1. Create all folders (with new name if renamed)
    # 2.2. Compile a list of files to be copied and a list of files to be moved (with new name recorded if renamed)
    list_copy_files = []
    list_move_files = []
    dataset_structure = soda["dataset-structure"]

    for folder_key, folder in dataset_structure["folders"].items():
        folderpath = join(datasetpath, folder_key)
        mkdir(folderpath)
        list_copy_files, list_move_files = recursive_dataset_scan(
            folder, folderpath, list_copy_files, list_move_files
        )

    # 3. Add high-level metadata files in the list
    if "dataset_metadata" in soda.keys():
        logger.info("generate_dataset_locally (optional) step 3 handling dataset_metadata")
        metadata_files = soda["dataset_metadata"]
        # log the metadata files that will be created
        for file_key, _ in metadata_files.items():
            if file_key == "subjects":
                subjects.create_excel(soda, False, join(datasetpath, "subjects.xlsx"))
            elif file_key == "samples":
                samples.create_excel(soda, False, join(datasetpath, "samples.xlsx"))
            elif file_key == "code_description":
                code_description.create_excel(soda, False, join(datasetpath, "code_description.xlsx"))
            elif file_key == "dataset_description": 
                dataset_description.create_excel(soda, False, join(datasetpath, "dataset_description.xlsx"))
            elif file_key == "performances":
                performances.create_excel(soda, False, join(datasetpath, "performances.xlsx"))
            elif file_key == "resources":
                resources.create_excel(soda, False, join(datasetpath, "resources.xlsx"))
            elif file_key == "sites":
                sites.create_excel(soda, False, join(datasetpath, "sites.xlsx"))
            elif file_key == "submission":
                submission.create_excel(soda, False, join(datasetpath, "submission.xlsx"))
            elif file_key == "README.md":
                text_metadata.create_text_file(soda, False, join(datasetpath, "README.md"), "README.md")
            elif file_key == "CHANGES":
                text_metadata.create_text_file(soda, False, join(datasetpath, "CHANGES"), "CHANGES")
            elif file_key == "LICENSE":
                text_metadata.create_text_file(soda, False, join(datasetpath, "LICENSE"), "LICENSE")

    # 4. Add manifest files in the list
    if "manifest_file" in soda["dataset_metadata"].keys():
        logger.info("generate_dataset_locally (optional) step 4 handling manifest-files")
        main_curate_progress_message = "Preparing manifest files"
        manifest.create_excel(soda, False, join(datasetpath,  "manifest.xlsx"))


    logger.info("generate_dataset_locally step 5 moving files to new location")
    # 5. Move files to new location
    main_curate_progress_message = "Moving files to new location"
    for fileinfo in list_move_files:
        srcfile = fileinfo[0]
        distfile = fileinfo[1]
        main_curate_progress_message = f"Moving file {str(srcfile)} to {str(distfile)}"
        shutil.move(srcfile, distfile)

    logger.info("generate_dataset_locally step 6 copying files to new location")
    # 6. Copy files to new location
    main_curate_progress_message = "Copying files to new location"
    start_generate = 1
    for fileinfo in list_copy_files:
        srcfile = fileinfo[0]
        distfile = fileinfo[1]
        main_curate_progress_message = f"Copying file {str(srcfile)} to {str(distfile)}"
        # track amount of copied files for loggin purposes
        mycopyfile_with_metadata(srcfile, distfile)
        main_curation_uploaded_files += 1

    logger.info("generate_dataset_locally step 7")
    # 7. Delete manifest folder and original folder if merge requested and rename new folder
    shutil.rmtree(manifest_folder_path) if isdir(manifest_folder_path) else 0
    if if_existing == "merge":
        logger.info("generate_dataset_locally (optional) step 7.1 delete manifest folder if merge requested")
        main_curate_progress_message = "Finalizing dataset"
        original_dataset_path = join(dataset_absolute_path, dataset_name)
        shutil.rmtree(original_dataset_path)
        rename(datasetpath, original_dataset_path)
        open_file(join(dataset_absolute_path, original_dataset_path))
    else:
        open_file(join(dataset_absolute_path, datasetpath))
    return datasetpath, main_total_generate_dataset_size


    


def ps_create_new_dataset(datasetname, ps):
    """
    Args:
        datasetname: name of the dataset to be created (string)
        bf: Pennsieve account object
    Action:
        Creates dataset for the account specified
    """
    try:
        error, count = "", 0
        datasetname = datasetname.strip()

        if check_forbidden_characters_ps(datasetname):
            error = (
                f"{error}Error: A Pennsieve dataset name cannot contain any of the following characters: "
                + forbidden_characters_bf
                + "<br>"
            )
            count += 1

        if not datasetname:
            error = f"{error}Error: Please enter valid dataset name<br>"
            count += 1

        if datasetname.isspace():
            error = error + "Error: Please enter valid dataset name" + "<br>"
            count += 1

        if count > 0:
            raise PennsieveDatasetNameInvalid(datasetname)

        try:
            dataset_list = get_users_dataset_list()
        except Exception as e:
            raise Exception("Failed to retrieve datasets from Pennsieve. Please try again later.")

        for dataset in dataset_list:
            if datasetname == dataset["content"]["name"]:
                raise PennsieveDatasetNameTaken("Dataset name already exists")
            
        
        # Create the dataset on Pennsieve
        r = requests.post(f"{PENNSIEVE_URL}/datasets", headers=create_request_headers(get_access_token()), json={"name": datasetname})
        r.raise_for_status()


        return r.json()

    # TODO: Remove unnecessary raise
    except Exception as e:
        raise e

double_extensions = [
    ".ome.tiff",
    ".ome.tif",
    ".ome.tf2,",
    ".ome.tf8",
    ".ome.btf",
    ".ome.xml",
    ".brukertiff.gz",
    ".mefd.gz",
    ".moberg.gz",
    ".nii.gz",
    ".mgh.gz",
    ".tar.gz",
    ".bcl.gz",
]


def create_high_lvl_manifest_files_existing_ps(
    soda, ps, my_tracking_folder
):
    """
    Function to create manifest files for each high-level SPARC folder.

    Args:
        soda: soda dict with information about the dataset to be generated/modified
    Action:
        manifest_files_structure: dict including the local path of the manifest files
    """
    def get_name_extension(file_name):
        double_ext = False
        for ext in double_extensions:
            if file_name.find(ext) != -1:
                double_ext = True
                break
        ext = ""
        name = ""
        if double_ext == False:
            name = os.path.splitext(file_name)[0]
            ext = os.path.splitext(file_name)[1]
        else:
            ext = (
                os.path.splitext(os.path.splitext(file_name)[0])[1]
                + os.path.splitext(file_name)[1]
            )
            name = os.path.splitext(os.path.splitext(file_name)[0])[0]
        return name, ext

    def recursive_import_ps_manifest_info(
        folder, my_relative_path, dict_folder_manifest, manifest_df
    ):
        """
        Import manifest information from the Pennsieve dataset for the given folder and its children.
        """

        if len(folder['children']) == 0:
            limit = 100
            offset = 0 
            ps_folder = {"children": []}
            while True: 
                r = requests.get(f"{PENNSIEVE_URL}/packages/{folder['content']['id']}?limit={limit}&offset={offset}", headers=create_request_headers(get_access_token()), json={"include": "files"})
                r.raise_for_status()
                page = r.json()
                normalize_tracking_folder(page)
                ps_folder["children"].extend(page)

                if len(page) < limit:
                    break
                offset += limit

            folder['children'] = ps_folder['children']

        for _, folder_item in folder["children"]["folders"].items():
            folder_name = folder_item['content']['name']
            relative_path = generate_relative_path(
                my_relative_path, folder_name
            )
            dict_folder_manifest = recursive_import_ps_manifest_info(
                folder_item, relative_path, dict_folder_manifest, manifest_df
            )
        for _, file in folder["children"]["files"].items():
            if file['content']['name'] != "manifest":
                file_id = file['content']['id']
                r = requests.get(f"{PENNSIEVE_URL}/packages/{file_id}/view", headers=create_request_headers(get_access_token()))
                r.raise_for_status()
                file_details = r.json()
                file_name = file_details[0]["content"]["name"]
                file_extension = splitext(file_name)[1]
                file_name_with_extension = (
                    splitext(file['content']['name'])[0] + file_extension
                )
                relative_path = generate_relative_path(
                    my_relative_path, file_name_with_extension
                )
                dict_folder_manifest["filename"].append(relative_path)
                # file type
                file_extension = get_name_extension(file_name)
                if file_extension == "":
                    file_extension = "None"
                dict_folder_manifest["file type"].append(file_extension)
                # timestamp, description, Additional Metadata
                if not manifest_df.empty:
                    if relative_path in manifest_df["filename"].values:
                        timestamp = manifest_df[
                            manifest_df["filename"] == relative_path
                        ]["timestamp"].iloc[0]
                        description = manifest_df[
                            manifest_df["filename"] == relative_path
                        ]["description"].iloc[0]
                        additional_metadata = manifest_df[
                            manifest_df["filename"] == relative_path
                        ]["Additional Metadata"].iloc[0]
                    else:
                        timestamp = ""
                        description = ""
                        additional_metadata = ""
                    dict_folder_manifest["timestamp"].append(timestamp)
                    dict_folder_manifest["description"].append(description)
                    dict_folder_manifest["Additional Metadata"].append(
                        additional_metadata
                    )
                else:
                    dict_folder_manifest["timestamp"].append("")
                    dict_folder_manifest["description"].append("")
                    dict_folder_manifest["Additional Metadata"].append("")
        return dict_folder_manifest

    # Merge existing folders
    def recursive_manifest_builder_existing_ps(
        my_folder,
        my_bf_folder,
        my_bf_folder_exists,
        my_relative_path,
        dict_folder_manifest,
    ):
        if "folders" in my_folder.keys():
            if my_bf_folder_exists:
                (
                    my_bf_existing_folders_name,
                ) = ps_get_existing_folders_details(my_bf_folder['children']['folders'])
            else:
                my_bf_existing_folders_name = []
            for folder_key, folder in my_folder["folders"].items():
                relative_path = generate_relative_path(my_relative_path, folder_key)
                if folder_key in my_bf_existing_folders_name:
                    bf_folder = my_bf_folder["children"]["folders"][folder_key]
                    bf_folder_exists = True
                else:
                    bf_folder = ""
                    bf_folder_exists = False
                dict_folder_manifest = recursive_manifest_builder_existing_ps(
                    folder,
                    bf_folder,
                    bf_folder_exists,
                    relative_path,
                    dict_folder_manifest,
                )
        if "files" in my_folder.keys():
            if my_bf_folder_exists:
                (
                    my_bf_existing_files_name,
                    my_bf_existing_files_name_with_extension,
                ) = ps_get_existing_files_details(my_bf_folder)
            else:
                my_bf_existing_files = []
                my_bf_existing_files_name = []
                my_bf_existing_files_name_with_extension = []
            for file_key, file in my_folder["files"].items():
                if file.get("location") == "local":
                    file_path = file["path"]
                    if isfile(file_path):
                        desired_name = splitext(file_key)[0]
                        file_extension = splitext(file_key)[1]
                        # manage existing file request
                        if existing_file_option == "skip" and file_key in my_bf_existing_files_name_with_extension:
                            continue
                        if existing_file_option == "replace" and file_key in my_bf_existing_files_name_with_extension:
                            # remove existing from manifest
                            filename = generate_relative_path(
                                my_relative_path, file_key
                            )
                            filename_list = dict_folder_manifest["filename"]
                            index_file = filename_list.index(filename)
                            del dict_folder_manifest["filename"][index_file]
                            del dict_folder_manifest["timestamp"][index_file]
                            del dict_folder_manifest["description"][index_file]
                            del dict_folder_manifest["file type"][index_file]
                            del dict_folder_manifest["Additional Metadata"][
                                index_file
                            ]
                            index_name = (
                                my_bf_existing_files_name_with_extension.index(
                                    file_key
                                )
                            )
                            del my_bf_existing_files[index_name]
                            del my_bf_existing_files_name[index_name]
                            del my_bf_existing_files_name_with_extension[
                                index_name
                            ]
                        if desired_name not in my_bf_existing_files_name:
                            final_name = file_key
                        else:
                            # expected final name
                            count_done = 0
                            final_name = desired_name
                            output = get_base_file_name(desired_name)
                            if output:
                                base_name = output[0]
                                count_exist = output[1]
                                while count_done == 0:
                                    if final_name in my_bf_existing_files_name:
                                        count_exist += 1
                                        final_name = (
                                            base_name + "(" + str(count_exist) + ")"
                                        )
                                    else:
                                        count_done = 1
                            else:
                                count_exist = 0
                                while count_done == 0:
                                    if final_name in my_bf_existing_files_name:
                                        count_exist += 1
                                        final_name = (
                                            desired_name
                                            + " ("
                                            + str(count_exist)
                                            + ")"
                                        )
                                    else:
                                        count_done = 1
                            final_name = final_name + file_extension
                            my_bf_existing_files_name.append(
                                splitext(final_name)[0]
                            )
                        # filename
                        filename = generate_relative_path(
                            my_relative_path, final_name
                        )
                        dict_folder_manifest["filename"].append(filename)
                        # timestamp
                        file_path = file["path"]
                        filepath = pathlib.Path(file_path)
                        mtime = filepath.stat().st_mtime
                        lastmodtime = datetime.fromtimestamp(mtime).astimezone(
                            local_timezone
                        )
                        dict_folder_manifest["timestamp"].append(
                            lastmodtime.isoformat()
                            .replace(".", ",")
                            .replace("+00:00", "Z")
                        )
                        # description
                        if "description" in file.keys():
                            dict_folder_manifest["description"].append(
                                file["description"]
                            )
                        else:
                            dict_folder_manifest["description"].append("")
                        # file type
                        if file_extension == "":
                            file_extension = "None"
                        dict_folder_manifest["file type"].append(file_extension)
                        # addtional metadata
                        if "additional-metadata" in file.keys():
                            dict_folder_manifest["Additional Metadata"].append(
                                file["additional-metadata"]
                            )
                        else:
                            dict_folder_manifest["Additional Metadata"].append("")
        return dict_folder_manifest

    double_extensions = [
        ".ome.tiff",
        ".ome.tif",
        ".ome.tf2,",
        ".ome.tf8",
        ".ome.btf",
        ".ome.xml",
        ".brukertiff.gz",
        ".mefd.gz",
        ".moberg.gz",
        ".nii.gz",
        ".mgh.gz",
        ".tar.gz",
        ".bcl.gz",
    ]

    try:
        # create local folder to save manifest files temporarly (delete any existing one first)
        shutil.rmtree(manifest_folder_path) if isdir(manifest_folder_path) else 0
        makedirs(manifest_folder_path)

        # import info about files already on ps
        dataset_structure = soda["dataset-structure"]
        manifest_dict_save = {}
        for high_level_folder_key, high_level_folder in my_tracking_folder["children"]["folders"].items():
            if (
                high_level_folder_key in dataset_structure["folders"].keys()
            ):

                relative_path = ""
                dict_folder_manifest = {}
                # Initialize dict where manifest info will be stored
                dict_folder_manifest["filename"] = []
                dict_folder_manifest["timestamp"] = []
                dict_folder_manifest["description"] = []
                dict_folder_manifest["file type"] = []
                dict_folder_manifest["Additional Metadata"] = []

                # pull manifest file into if exists 
                manifest_df = pd.DataFrame()
                for file_key, file in high_level_folder['children']['files'].items():
                    file_id = file['content']['id']
                    r = requests.get(f"{PENNSIEVE_URL}/packages/{file_id}/view", headers=create_request_headers(get_access_token()))
                    r.raise_for_status()
                    file_details = r.json()
                    file_name_with_extension = file_details[0]["content"]["name"]
                    if file_name_with_extension in manifest_sparc:
                        file_id_2 = file_details[0]["content"]["id"]
                        r = requests.get(f"{PENNSIEVE_URL}/packages/{file_id}/files/{file_id_2}", headers=create_request_headers(get_access_token()))
                        r.raise_for_status()
                        file_url_info = r.json()
                        file_url = file_url_info["url"]
                        manifest_df = pd.read_excel(file_url, engine="openpyxl")
                        manifest_df = manifest_df.fillna("")
                        if (
                            "filename" not in manifest_df.columns
                            or "description" not in manifest_df.columns
                            or "Additional Metadata" not in manifest_df.columns
                        ):
                            manifest_df = pd.DataFrame()
                        break

                # store the data frame pulled from Pennsieve into a dictionary
                dict_folder_manifest =  recursive_import_ps_manifest_info(
                    high_level_folder, relative_path, dict_folder_manifest, manifest_df
                )

                manifest_dict_save[high_level_folder_key] = {
                    "manifest": dict_folder_manifest,
                    "bf_folder": high_level_folder,
                }

        # import info from local files to be uploaded
        local_timezone = TZLOCAL()
        manifest_files_structure = {}
        existing_folder_option = soda["generate-dataset"]["if-existing"]
        existing_file_option = soda["generate-dataset"][
            "if-existing-files"
        ]
        for folder_key, folder in dataset_structure["folders"].items():
            relative_path = ""

            if (
                folder_key in manifest_dict_save
                and existing_folder_option == "merge"
            ):
                bf_folder = manifest_dict_save[folder_key]["bf_folder"]
                bf_folder_exists = True
                dict_folder_manifest = manifest_dict_save[folder_key]["manifest"]

            elif (
                folder_key in manifest_dict_save
                and folder_key
                not in my_tracking_folder["children"]["folders"].keys()
                and existing_folder_option == "skip"
            ):
                continue

            else:
                bf_folder = ""
                bf_folder_exists = False
                dict_folder_manifest = {}
                dict_folder_manifest["filename"] = []
                dict_folder_manifest["timestamp"] = []
                dict_folder_manifest["description"] = []
                dict_folder_manifest["file type"] = []
                dict_folder_manifest["Additional Metadata"] = []

            dict_folder_manifest = recursive_manifest_builder_existing_ps(
                folder, bf_folder, bf_folder_exists, relative_path, dict_folder_manifest
            )

            # create high-level folder at the temporary location
            folderpath = join(manifest_folder_path, folder_key)
            makedirs(folderpath)

            # save manifest file
            manifestfilepath = join(folderpath, "manifest.xlsx")
            df = pd.DataFrame.from_dict(dict_folder_manifest)
            df.to_excel(manifestfilepath, index=None, header=True)
            wb = load_workbook(manifestfilepath)
            ws = wb.active

            blueFill = PatternFill(
                start_color="9DC3E6", fill_type="solid"
            )
            greenFill = PatternFill(
                start_color="A8D08D", fill_type="solid"
            )
            yellowFill = PatternFill(
                start_color="FFD965", fill_type="solid"
            )
            ws['A1'].fill = blueFill
            ws['B1'].fill = greenFill
            ws['C1'].fill = greenFill
            ws['D1'].fill = greenFill
            ws['E1'].fill = yellowFill
            wb.save(manifestfilepath)

            manifest_files_structure[folder_key] = manifestfilepath

        return manifest_files_structure

    except Exception as e:
        raise e





def generate_relative_path(x, y):
    return x + "/" + y if x else y


def ps_get_existing_folders_details(ps_folders):
    ps_existing_folders = [ps_folders[folder] for folder in ps_folders if ps_folders[folder]["content"]["packageType"] == "Collection"]
    ps_existing_folders_name = [folder['content']["name"] for folder in ps_existing_folders]

    return ps_existing_folders, ps_existing_folders_name


def ps_get_existing_files_details(ps_folder):
    # TODO: Dorian -> ["extensions doesn't seem to be returned anymore by the endpoint"]
    def verify_file_name(file_name, extension):
        if extension == "":
            return file_name

        double_ext = False
        for ext in double_extensions:
            if file_name.find(ext) != -1:
                double_ext = True
                break

        extension_from_name = ""

        if double_ext == False:
            extension_from_name = os.path.splitext(file_name)[1]
        else:
            extension_from_name = (
                os.path.splitext(os.path.splitext(file_name)[0])[1]
                + os.path.splitext(file_name)[1]
            )

        if extension_from_name == ("." + extension):
            return file_name
        else:
            return file_name + ("." + extension)

    files = ps_folder["children"]["files"]
    double_extensions = [
        ".ome.tiff",
        ".ome.tif",
        ".ome.tf2,",
        ".ome.tf8",
        ".ome.btf",
        ".ome.xml",
        ".brukertiff.gz",
        ".mefd.gz",
        ".moberg.gz",
        ".nii.gz",
        ".mgh.gz",
        ".tar.gz",
        ".bcl.gz",
    ]


    bf_existing_files_name = [splitext(files[file]['content']["name"])[0] for file in files]
    bf_existing_files_name_with_extension = []

    # determine if we are at the root of the dataset
    content = ps_folder["content"]
    if (str(content['id'])[2:9]) == "dataset":
        r = requests.get(f"{PENNSIEVE_URL}/datasets/{content['id']}", headers=create_request_headers(get_access_token())) 
        r.raise_for_status()
        root_folder = r.json()
        root_children = root_folder["children"]
        for item in root_children:
            file_name_with_extension = ""
            item_id = item["content"]["id"]
            item_name = item["content"]["name"]
            if item_id[2:9] == "package":
                if("extension" not in root_children):
                    file_name_with_extension = verify_file_name(item_name,"")
                else:
                    file_name_with_extension = verify_file_name(item_name, root_children["extension"])

            if file_name_with_extension == "":
                continue
            bf_existing_files_name_with_extension.append(file_name_with_extension)
    else:
        #is collection - aka a folder in the dataset
        for file_key, file in files.items():
            file_name_with_extension = ""
            file_name = file["content"]["name"]
            file_id = file["content"]["id"]
            if file_id[2:9] == "package":
                if "extension" not in file:
                    file_name_with_extension = verify_file_name(file_name,"")
                else:
                    file_name_with_extension = verify_file_name(file_name, file["extension"])
            if file_name_with_extension == "":
                continue
            bf_existing_files_name_with_extension.append(file_name_with_extension)


    return (
        bf_existing_files_name,
        bf_existing_files_name_with_extension,
    )


def check_if_int(s):
    try:
        int(s)
        return True
    except ValueError:
        return False


def get_base_file_name(file_name):
    output = []
    if file_name[-1] == ")":
        string_length = len(file_name)
        count_start = string_length
        character = file_name[count_start - 1]
        while character != "(" and count_start >= 0:
            count_start -= 1
            character = file_name[count_start - 1]
        if character == "(":
            base_name = file_name[:count_start - 1]
            num = file_name[count_start : string_length - 1]
            if check_if_int(num):
                output = [base_name, int(num)]
    return output


def clean_existing_ps_dataset(soda, ds):
    global logger

    logger.info("Starting ps_update_existing_dataset")

    global main_curate_progress_message
    global main_total_generate_dataset_size
    global start_generate
    global main_initial_bfdataset_size

    # Delete any files on Pennsieve that have been marked as deleted
    def recursive_file_delete(folder):
        if "files" in folder.keys():
            for item in list(folder["files"]):
                if "deleted" in folder["files"][item]["action"]:
                    file_path = folder["files"][item]["path"]
                    # remove the file from the dataset
                    r = requests.post(f"{PENNSIEVE_URL}/data/delete", headers=create_request_headers(get_access_token()), json={"things": [file_path]})
                    r.raise_for_status()
                    # remove the file from the soda json structure
                    del folder["files"][item]

        for item in list(folder["folders"]):
            recursive_file_delete(folder["folders"][item])

    def recursive_item_path_create(folder, path):
        """
        Recursively create the path for the item    # Add a new key containing the path to all the files and folders on the
        local data structure.
        Allows us to see if the folder path of a specfic file already
        exists on Pennsieve.
        """
        
        if "files" in folder.keys():
            for item in list(folder["files"]):
                if item in ["manifest.xslx", "manifest.csv"]:
                    continue
                if "folderpath" not in folder["files"][item]:
                    folder["files"][item]["folderpath"] = path[:]

        if "folders" in folder.keys():
            for item in list(folder["folders"]):
                if "folderpath" not in folder["folders"][item]:
                    folder["folders"][item]["folderpath"] = path[:]
                    folder["folders"][item]["folderpath"].append(item)
                recursive_item_path_create(
                    folder["folders"][item], folder["folders"][item]["folderpath"][:]
                )

        return

    # Check and create any non existing folders for the file move process (Used in the recursive_check_moved_files function)
    def recursive_check_and_create_ps_file_path(
        folderpath, index, current_folder_structure
    ):
        folder = folderpath[index]

        if folder not in current_folder_structure["folders"]:
            if index == 0:
                r = requests.post(f"{PENNSIEVE_URL}/packages", json={"name": folder, "parent": f"{current_folder_structure['path']}", "packageType": "collection", "dataset": ds['content']['id']},  headers=create_request_headers(get_access_token()))
                r.raise_for_status()
                new_folder = r.json()
            else:
                r = requests.post(f"{PENNSIEVE_URL}/packages", json={"name": folder, "parent": f"{current_folder_structure['path']}", "packageType": "collection", "dataset": ds['content']['id']},  headers=create_request_headers(get_access_token()))
                r.raise_for_status()
                new_folder = r.json()
            
            current_folder_structure["folders"][folder] = {
                "location": "ps",
                "action": ["existing"],
                "path": new_folder['content']['id'],
                "folders": {},
                "files": {},
            }

        index += 1
        # check if path exists for folder, if not then folder has not been created on Pennsieve yet, so create it and add it to the path key
        if "path" not in current_folder_structure["folders"][folder].keys() or current_folder_structure["folders"][folder]["location"] != "ps":
            r = requests.post(f"{PENNSIEVE_URL}/packages", headers=create_request_headers(get_access_token()), json=build_create_folder_request(folder, current_folder_structure["path"], ds['content']['id']))
            r.raise_for_status()
            new_folder_id = r.json()["content"]["id"]
            current_folder_structure["folders"][folder]["path"] = new_folder_id

        if index < len(folderpath):
            return recursive_check_and_create_ps_file_path(
                folderpath, index, current_folder_structure["folders"][folder]
            )
        else:
            return current_folder_structure["folders"][folder]["path"]


    # Rename any files that exist on Pennsieve
    def recursive_file_rename(folder):
        if "files" in folder.keys():
            for item in list(folder["files"]):
                if (
                    "renamed" in folder["files"][item]["action"]
                    and folder["files"][item]["location"] == "ps"
                ):
                    # rename the file on Pennsieve
                    r = requests.put(f"{PENNSIEVE_URL}/packages/{folder['files'][item]['path']}?updateStorage=true", json={"name": item}, headers=create_request_headers(get_access_token()))
                    r.raise_for_status()

        for item in list(folder["folders"]):
            recursive_file_rename(folder["folders"][item])


    def recursive_folder_delete(folder):
        """
        Delete any stray folders that exist on Pennsieve
        Only top level files are deleted since the api deletes any
        files and folders that exist inside.
        """
        for item in list(folder["folders"]):
            child = folder["folders"][item]
            if child.get("location") == "ps":
                if "moved" in child.get("action", []):
                    file_path = child["path"]
                    # remove the file from the dataset
                    r = requests.post(f"{PENNSIEVE_URL}/data/delete", headers=create_request_headers(get_access_token()), json={"things": [file_path]})
                    r.raise_for_status()
                if "deleted" in child.get("action", []):
                    file_path = child["path"]
                    # remove the file from the dataset
                    r = requests.post(f"{PENNSIEVE_URL}/data/delete", headers=create_request_headers(get_access_token()), json={"things": [file_path]})
                    r.raise_for_status()
                    del folder["folders"][item]
                else:
                    recursive_folder_delete(child)
            else:
                recursive_folder_delete(child)


    # Rename any folders that still exist.
    def recursive_folder_rename(folder, mode):
        for item in list(folder["folders"]):
            child = folder["folders"][item]
            if (
                child.get("location") == "ps"
                and "action" in child
                and mode in child["action"]
            ):
                folder_id = child["path"]
                r = requests.put(f"{PENNSIEVE_URL}/packages/{folder_id}?updateStorage=true", headers=create_request_headers(get_access_token()), json={"name": item})
                r.raise_for_status()
            recursive_folder_rename(child, mode)


    ps_dataset = ""
    start = timer()
    # 1. Remove all existing files on Pennsieve, that the user deleted.
    logger.info("ps_update_existing_dataset step 1 remove existing files on Pennsieve the user deleted")
    main_curate_progress_message = "Checking Pennsieve for deleted files"
    dataset_structure = soda["dataset-structure"]
    recursive_file_delete(dataset_structure)
    main_curate_progress_message = (
        "Files on Pennsieve marked for deletion have been deleted"
    )

    # 2. Rename any deleted folders on Pennsieve to allow for replacements.
    logger.info("ps_update_existing_dataset step 2 rename deleted folders on Pennsieve to allow for replacements")
    main_curate_progress_message = "Checking Pennsieve for deleted folders"
    dataset_structure = soda["dataset-structure"]
    recursive_folder_rename(dataset_structure, "deleted")
    main_curate_progress_message = "Folders on Pennsieve have been marked for deletion"

    # 2.5 Rename folders that need to be in the final destination.
    logger.info("ps_update_existing_dataset step 2.5 rename folders that need to be in the final destination")
    main_curate_progress_message = "Renaming any folders requested by the user"
    recursive_folder_rename(dataset_structure, "renamed")
    main_curate_progress_message = "Renamed all folders requested by the user"

    # 3. Get the status of all files currently on Pennsieve and create
    # the folderpath for all items in both dataset structures.
    logger.info("ps_update_existing_dataset step 3 get the status of all files currently on Pennsieve and create the folderpath for all items in both dataset structures")
    main_curate_progress_message = "Fetching files and folders from Pennsieve"
    current_bf_dataset_files_folders = import_pennsieve_dataset(
        soda.copy()
    )["soda_object"]
    ps_dataset = current_bf_dataset_files_folders["dataset-structure"]
    main_curate_progress_message = "Creating file paths for all files on Pennsieve"
    recursive_item_path_create(dataset_structure, [])
    recursive_item_path_create(ps_dataset, [])
    main_curate_progress_message = "File paths created"

    # 4. Run the original code to upload any new files added to the dataset.
    logger.info("ps_update_existing_dataset step 4 run the ps_create_new_dataset code to upload any new files added to the dataset")
    if "dataset_metadata" in soda.keys() and "manifest_files" in soda["dataset_metadata"].keys():
        if "auto-generated" in soda["manifest-files"].keys():
            soda["manifest-files"] = {"destination": "ps", "auto-generated": True}
        else:
            soda["manifest-files"] = {"destination": "ps"}

    end = timer()
    logger.info(f"Time for clean_existing_ps_dataset: {timedelta(seconds=end - start)}")


def get_origin_manifest_id(dataset_id):
    global logger
    max_attempts = 3
    for _ in range(max_attempts):
        manifests = get_upload_manifests(dataset_id)
        if manifests and "manifests" in manifests and manifests["manifests"]:
            # sort the manifests list by date_created timestamp field in descending order
            manifests["manifests"].sort(key=lambda x: x["date_created"], reverse=True)
            return manifests["manifests"][0]["id"]
        time.sleep(5)  # Wait for 5 seconds before the next attempt

    raise Exception("Did not get the origin manifest id in an expected amount of time.")



def normalize_tracking_folder(tracking_folder):
    """
    Normalize the tracking folder object to be a dictonary with the shape: {files: {}, folders: {}}. 
    This shape matches our dataset structure object. Recall, the tracking folder receives information about what folders and 
    files are stored on Pennsieve. We update this as we update Pennsieve's state. 
    """
    if tracking_folder == "":
        return {"folders": {}, "files": {} }
    
    temp_children = {"folders": {}, "files": {}}


    # add the files and folders to the temp_children structure 
    for child in tracking_folder["children"]:
        if child["content"]["packageType"] == "Collection":
            # add the folders ( designated collection on Pennsieve ) to the temp_children structure under folders
            temp_children["folders"][child["content"]["name"]] = child
        else:
            # add the files (anything not designated a collection) to the temp_children structure under files
            temp_children["files"][child["content"]["name"]] = child

    # replace the non-normalized children structure with the normalized children structure
    tracking_folder["children"] = temp_children


def build_create_folder_request(folder_name, folder_parent_id, dataset_id):
    """
    Create a folder on Pennsieve. 
    """
    body = {}

    # if creating a folder at the root of the dataset the api does not require a parent key
    if folder_parent_id.find("N:dataset") == -1:
        body["parent"] = folder_parent_id
    
    body["name"] = folder_name
    body["dataset"] = dataset_id
    body["packageType"] = "collection"

    return body


bytes_uploaded_per_file = {}
total_bytes_uploaded = {"value": 0}
current_files_in_subscriber_session = 0



bytes_file_path_dict = {}

# retry variables instantiated outside function
list_of_files_to_rename = {}
renamed_files_counter = 0
total_files = 0
total_metadata_files = 0
total_manifest_files = 0


def create_metadata_files_for_upload(soda, list_upload_metadata_files, existing_root_files=None, existing_file_option="skip", ps=None):
    """
    Creates metadata files (Excel and text) based on soda["dataset_metadata"] and appends them to the upload list.
    Updates global counters for total files and sizes.
    
    Args:
        soda: The soda configuration object containing dataset_metadata
        list_upload_metadata_files: List to append created metadata file paths to
        existing_root_files: Dict of files already at the root of the Pennsieve dataset (optional)
        existing_file_option: How to handle existing files - "skip", "replace", or "merge" (default: "skip")
        ps: Pennsieve client object (required for delete operations when using "replace")
    """
    global main_total_generate_dataset_size
    global total_files
    global total_metadata_files

    logger.info(f"create_metadata_files_for_upload: Starting with existing_file_option='{existing_file_option}', existing_root_files={'present' if existing_root_files else 'None'}")

    if "dataset_metadata" not in soda or soda["dataset_metadata"] == {}:
        logger.info("create_metadata_files_for_upload: No dataset_metadata found in soda, returning early")
        return

    # Normalize existing files for lookup (case-insensitive filename -> file info)
    existing_files_map = {}
    if existing_root_files:
        for file_key, file_info in existing_root_files.items():
            existing_files_map[file_key.lower()] = file_info
        logger.info(f"create_metadata_files_for_upload: Found {len(existing_files_map)} existing files at dataset root: {list(existing_files_map.keys())}")

    files_created = 0
    files_skipped = 0
    files_deleted = 0

    def delete_existing_file(filename):
        """Delete an existing file on Pennsieve if it exists and we have the ps client."""
        nonlocal files_deleted
        filename_lower = filename.lower()
        if filename_lower not in existing_files_map:
            return
        
        file_info = existing_files_map[filename_lower]
        file_id = file_info.get("content", {}).get("id")
        
        if not file_id:
            logger.warning(f"create_metadata_files_for_upload: Could not get file ID for '{filename}' to delete")
            return
        
        if not ps:
            logger.warning(f"create_metadata_files_for_upload: Cannot delete '{filename}' - ps client not provided")
            return
        
        try:
            r = requests.post(f"{PENNSIEVE_URL}/data/delete", headers=create_request_headers(get_access_token()), json={"things": [file_id]})
            r.raise_for_status()
            files_deleted += 1
            logger.info(f"create_metadata_files_for_upload: Deleted existing '{filename}' (ID: {file_id}) on Pennsieve")
        except Exception as e:
            logger.error(f"create_metadata_files_for_upload: Failed to delete '{filename}': {e}")

    def should_upload_file(filename):
        """Check if a metadata file should be uploaded based on existing files and option."""
        nonlocal files_skipped
        if not existing_root_files:
            return True  # No existing files to check against
        
        filename_lower = filename.lower()
        file_exists = filename_lower in existing_files_map
        
        if file_exists and existing_file_option == "skip":
            logger.info(f"create_metadata_files_for_upload: Skipping '{filename}' - already exists on Pennsieve (if-existing-files='skip')")
            files_skipped += 1
            return False
        
        return True  # File doesn't exist or will be replaced/merged

    def add_metadata_file(filepath, filename):
        """Helper to register a created metadata file and update counters."""
        nonlocal files_created
        global main_total_generate_dataset_size, total_files, total_metadata_files
        
        # If replacing, delete the existing file first
        if existing_file_option == "replace":
            delete_existing_file(filename)
        
        file_size = getsize(filepath)
        list_upload_metadata_files.append(filepath)
        main_total_generate_dataset_size += file_size
        total_files += 1
        total_metadata_files += 1
        files_created += 1
        logger.info(f"create_metadata_files_for_upload: Created '{filename}' ({file_size} bytes) - added to upload list")

    # Excel metadata: soda key -> (create function, output filename)
    EXCEL_METADATA = {
        "submission": (submission.create_excel, "submission.xlsx"),
        "subjects": (subjects.create_excel, "subjects.xlsx"),
        "samples": (samples.create_excel, "samples.xlsx"),
        "performances": (performances.create_excel, "performances.xlsx"),
        "resources": (resources.create_excel, "resources.xlsx"),
        "sites": (sites.create_excel, "sites.xlsx"),
        "dataset_description": (dataset_description.create_excel, "dataset_description.xlsx"),
        "code_description": (code_description.create_excel, "code_description.xlsx"),
        "manifest_file": (manifest.create_excel, "manifest.xlsx"),
    }

    # Text metadata files (soda key matches filename)
    TEXT_METADATA = ["README.md", "CHANGES", "LICENSE"]

    for key in soda["dataset_metadata"]:
        if key in EXCEL_METADATA:
            create_func, filename = EXCEL_METADATA[key]
            if not should_upload_file(filename):
                continue
            filepath = os.path.join(METADATA_UPLOAD_PS_PATH, filename)
            create_func(soda, False, filepath)
            add_metadata_file(filepath, filename)

        elif key in TEXT_METADATA:
            if not should_upload_file(key):
                continue
            filepath = os.path.join(METADATA_UPLOAD_PS_PATH, key)
            text_metadata.create_text_file(soda, False, filepath, key)
            add_metadata_file(filepath, key)

    logger.info(f"create_metadata_files_for_upload: Completed - {files_created} files created, {files_skipped} files skipped, {files_deleted} files deleted")




def create_upload_information_new(soda, ps, relative_path):
    list_upload_files = []
    list_upload_metadata_files = []
    main_total_generate_dataset_size = 0
    total_files = 0
    total_metadata_files = 0
    bytes_file_path_dict = {}

    global logger 
    logger.info("Create new manifest - helper function creating upload information new")

    def recursive_dataset_scan_for_new_upload(dataset_structure, list_upload_files, my_relative_path):
        """
        This function recursively gathers the files and folders in the dataset that will be uploaded to Pennsieve.
        It assumes the dataset is new based on the generate_option value and will spend less time comparing what is on Pennsieve.
        It will gather all the relative paths for the files and folders to pass along to the Pennsieve agent.
        Input:
        dataset_structure,
        my_relative_path

        Output:
        two lists in one tuple, the first list will have all the local file paths that will be uploaded to Pennsieve
        The second list will have the relative files paths according to the dataset structure.
        If the folder does not existing yet on Pennsieve the agent will create it.
        """
        nonlocal main_total_generate_dataset_size
        nonlocal bytes_file_path_dict
        # First loop will take place in the root of the dataset
        if "folders" in dataset_structure.keys():
            for folder_key, folder in dataset_structure["folders"].items():
                relative_path = generate_relative_path(my_relative_path, folder_key)
                list_upload_files = recursive_dataset_scan_for_new_upload(folder, list_upload_files, relative_path)
        if "files" in dataset_structure.keys():
            list_local_files = []
            list_projected_names = []
            list_desired_names = []
            list_final_names = []

            list_initial_names = []
            for file_key, file in dataset_structure["files"].items():
                # relative_path = generate_relative_path(my_relative_path, file_key)
                file_path = file["path"]
                if isfile(file_path) and file.get("location") == "local":
                    projected_name = splitext(basename(file_path))[0]
                    projected_name_w_extension = basename(file_path)
                    desired_name = splitext(file_key)[0]
                    desired_name_with_extension = file_key


                    if projected_name != desired_name:
                        list_initial_names.append(projected_name)
                        list_local_files.append(file_path)
                        list_projected_names.append(projected_name_w_extension)
                        list_desired_names.append(desired_name_with_extension)
                        list_final_names.append(desired_name)
                    else:
                        list_local_files.append(file_path)
                        list_projected_names.append(projected_name_w_extension)
                        list_desired_names.append(desired_name_with_extension)
                        list_final_names.append(desired_name)
                        list_initial_names.append(projected_name)

                    file_size = getsize(file_path)
                    main_total_generate_dataset_size += file_size
                    bytes_file_path_dict[file_path] = file_size

            if list_local_files:
                list_upload_files.append([
                    list_local_files,
                    list_projected_names,
                    list_desired_names,
                    list_final_names,
                    "/" if my_relative_path == soda["generate-dataset"]["dataset-name"] else my_relative_path,
                ])


        return list_upload_files


    # we can assume no files/folders exist in the dataset since the generate option is new and starting point is also new
    # therefore, we can assume the dataset structure is the same as the tracking structure

    list_upload_files = recursive_dataset_scan_for_new_upload(soda["dataset-structure"], list_upload_files, relative_path)

    if len(list_upload_files) <= 0:
        logger.info("Create new manifest - helper function failed to add new files to upload files") 

    create_metadata_files_for_upload(soda, list_upload_metadata_files, existing_file_option=False, existing_root_files=False, ps=ps )

    return {
        "list_upload_files": list_upload_files,
        "list_upload_metadata_files": list_upload_metadata_files,
        "total_files": total_files,
        "total_metadata_files": total_metadata_files,
        "main_total_generate_dataset_size": main_total_generate_dataset_size,
        "bytes_file_path_dict": bytes_file_path_dict
    }



    # Helper function


def create_upload_information_existing(soda, dataset_structure, ds, ps, relative_path):

    global main_curate_progress_message
    global main_curate_status


    # See how to create folders with the Pennsieve agent
    def recursive_create_folder_for_ps(
        my_folder, my_tracking_folder, existing_folder_option
    ):
        """
        Creates a folder on Pennsieve for each folder in the dataset structure if they aren't already present in the dataset.
        Input:
            my_folder: The dataset structure to be created on Pennsieve. Pass in the soda json object to start. 
            my_tracking_folder: Tracks what folders have been created on Pennsieve thus far. Starts as an empty dictionary.
            existing_folder_option: Dictates whether to merge, duplicate, replace, or skip existing folders.
        """
        # Check if the current folder has any subfolders that already exist on Pennsieve. Important step to appropriately handle replacing and merging folders.
        if len(my_tracking_folder["children"]["folders"]) == 0 and my_tracking_folder["content"]["id"].find("N:dataset") == -1:
            limit = 100
            offset = 0
            ps_folder = {}
            ps_folder_children = []
            while True: 
                r = requests.get(f"{PENNSIEVE_URL}/packages/{my_tracking_folder['content']['id']}?limit={limit}&offset={offset}", headers=create_request_headers(get_access_token()), json={"include": "files"})
                r.raise_for_status()
                ps_folder = r.json()
                page = ps_folder["children"]
                ps_folder_children.extend(page)
                if len(page) < limit:
                    break
                offset += limit
                time.sleep(1)
                
            ps_folder["children"] = ps_folder_children
            normalize_tracking_folder(ps_folder)
            my_tracking_folder["children"] = ps_folder["children"]

        # create/replace/skip folder
        if "folders" in my_folder.keys():
            for folder_key, folder in my_folder["folders"].items():
                if existing_folder_option == "merge":
                    if folder_key in my_tracking_folder["children"]["folders"]:
                        ps_folder = my_tracking_folder["children"]["folders"][folder_key]
                        normalize_tracking_folder(ps_folder)
                    else:
                        # We are merging but this is a new folder - not one that already exists in the current dataset - so we create it.
                        r = requests.post(f"{PENNSIEVE_URL}/packages", headers=create_request_headers(get_access_token()), json=build_create_folder_request(folder_key, my_tracking_folder['content']['id'], ds['content']['id']))
                        r.raise_for_status()
                        ps_folder = r.json()
                        normalize_tracking_folder(ps_folder)

                elif existing_folder_option == "replace":
                    # if the folder exists on Pennsieve remove it
                    if folder_key in my_tracking_folder["children"]["folders"]:
                        ps_folder = my_tracking_folder["children"]["folders"][folder_key]

                        r = requests.post(f"{PENNSIEVE_URL}/data/delete", headers=create_request_headers(get_access_token()), json={"things": [ps_folder["content"]["id"]]})
                        r.raise_for_status()

                        # remove from ps_folder 
                        del my_tracking_folder["children"]["folders"][folder_key]

                    r = requests.post(f"{PENNSIEVE_URL}/packages", headers=create_request_headers(get_access_token()), json=build_create_folder_request(folder_key, my_tracking_folder['content']['id'], ds['content']['id']))
                    r.raise_for_status()
                    ps_folder = r.json()
                    normalize_tracking_folder(ps_folder)

                my_tracking_folder["children"]["folders"][folder_key] = ps_folder
                tracking_folder = my_tracking_folder["children"]["folders"][folder_key] # get the folder we just added to the tracking folder
                recursive_create_folder_for_ps(
                    folder, tracking_folder, existing_folder_option
                )

    def recursive_dataset_scan_for_ps(
            my_folder,
            my_tracking_folder,
            existing_file_option,
            list_upload_files,
            my_relative_path,
        ):
            """
                Delete files that are marked to be replaced in the dataset. Create a list of files to upload to Pennsieve.
            """

            nonlocal main_total_generate_dataset_size
            global logger


            # folder children are packages such as collections and files stored on the Pennsieve dataset
            ps_folder_children = my_tracking_folder["children"] #ds (dataset)


            # Recursively go through the dataset structure and the Pennsieve dataset to find
            # any files that need to be uploaded, replaced, or skipped based on the existing_file_option.
            if "folders" in my_folder.keys():
                for folder_key, folder in my_folder["folders"].items():
                    relative_path = generate_relative_path(my_relative_path, folder_key)
                    tracking_folder = ps_folder_children["folders"][folder_key]
                    list_upload_files = recursive_dataset_scan_for_ps(
                        folder,
                        tracking_folder,
                        existing_file_option,
                        list_upload_files,
                        relative_path,
                    )

            if "files" in my_folder.keys(): 

                # delete files to be deleted
                (
                    my_bf_existing_files_name,
                    my_bf_existing_files_name_with_extension,
                ) = ps_get_existing_files_details(my_tracking_folder)

                for file_key, file in my_folder["files"].items():
                    # if local then we are either adding a new file to an existing/new dataset or replacing a file in an existing dataset
                    if file.get("location") == "local":
                        file_path = file["path"]
                        if isfile(file_path):
                            if ("renamed" in file.get("action", [])):
                                original_file_key = file.get("original-name", file_key)
                                if original_file_key in ps_folder_children["files"]:
                                    logger.info(f"list-upload-files log: Renaming file: Found original file '{original_file_key}' on Pennsieve (was renamed from '{file_key}'). Deleting it because it must be re-uploaded with the new name.")
                                    my_file = ps_folder_children["files"][original_file_key]
                                    # delete the package ( aka file ) from the dataset 
                                    r = requests.post(f"{PENNSIEVE_URL}/data/delete", headers=create_request_headers(get_access_token()), json={"things": [f"{my_file['content']['id']}"]})
                                    r.raise_for_status()
                                    del ps_folder_children["files"][original_file_key]
                            if file_key in ps_folder_children["files"] and existing_file_option == "replace":
                                logger.info(f"list-upload-files log: Found file '{file_key}' on Pennsieve for deletion")
                                my_file = ps_folder_children["files"][file_key]
                                # delete the package ( aka file ) from the dataset 
                                r = requests.post(f"{PENNSIEVE_URL}/data/delete", headers=create_request_headers(get_access_token()), json={"things": [f"{my_file['content']['id']}"]})
                                r.raise_for_status()
                                del ps_folder_children["files"][file_key]


                # create list of files to be uploaded with projected and desired names saved
                (
                    my_bf_existing_files_name,
                    my_bf_existing_files_name_with_extension,
                ) = ps_get_existing_files_details(my_tracking_folder)

                list_local_files = []
                list_projected_names = []
                list_desired_names = []
                list_final_names = []
                additional_upload_lists = []

                list_initial_names = []

                # add the files that are set to be uploaded to Pennsieve to a list 
                # handle renaming files and creating duplicates
                for file_key, file in my_folder["files"].items():
                    if file.get("location") == "local":
                        file_path = file["path"]
                        if isfile(file_path):
                            file_extension = file.get("extension", splitext(file_path)[1])

                            initial_name = None
                            initial_name_with_extension = None
                            desired_name_with_extension = None
                            if ("renamed" in file.get("action", [])):
                                # Remove the extension from the original name using file_extension
                                initial_name_with_extension = file.get("original-name", file_key)
                                if initial_name_with_extension.endswith(file_extension):
                                    initial_name = initial_name_with_extension[: -len(file_extension)]
                                else:
                                    initial_name = splitext(initial_name_with_extension)[0]
                                desired_name_with_extension = file_key
                            else:
                                #logic for non-renamed files
                                initial_name = file_key[: -len(file_extension)] if file_key.endswith(file_extension) else splitext(basename(file_path))[0]
                                initial_name_with_extension = basename(file_path)
                                desired_name_with_extension = file_key
                            
                            # Skip file if skip option is set and the desired name already exists on Pennsieve
                            if existing_file_option != "replace" and desired_name_with_extension in my_bf_existing_files_name_with_extension:
                                logger.info(f"list-upload-files log: File '{desired_name_with_extension}' already exists on Pennsieve and skip option is set - file will not be uploaded")
                                continue


                            logger.info(f"list-upload-files log: File '{file_key}' added to list_local_files with projected_name: '{initial_name_with_extension}', final_name: '{desired_name_with_extension}'")
                            list_local_files.append(file_path)
                            list_projected_names.append(initial_name_with_extension)
                            list_desired_names.append(desired_name_with_extension)
                            list_final_names.append(desired_name_with_extension)
                            list_initial_names.append(initial_name)

                            my_bf_existing_files_name_with_extension.append(desired_name_with_extension)

                            # add to projected dataset size to be generated
                            main_total_generate_dataset_size += getsize(file_path)

                if list_local_files:
                    ds_name = soda["ps-dataset-selected"]["dataset-name"]
                    list_upload_files.append(
                        [
                            list_local_files,
                            ps_folder_children,
                            list_projected_names,
                            list_desired_names,
                            list_final_names,
                            my_tracking_folder,
                            "/" if my_relative_path == ds_name else my_relative_path,
                        ]
                    )

                for item in additional_upload_lists:
                    list_upload_files.append(item)

            return list_upload_files



    list_upload_files = []
    list_upload_metadata_files = []
    list_upload_manifest_files = []
    total_files = 0
    total_metadata_files = 0
    total_manifest_files = 0
    main_total_generate_dataset_size = 0

    existing_folder_option = soda["generate-dataset"]["if-existing"]
    existing_file_option = soda["generate-dataset"][
        "if-existing-files"
    ]

    # we will need a tracking structure to compare against
    tracking_json_structure = ds
    normalize_tracking_folder(tracking_json_structure)
    recursive_create_folder_for_ps(dataset_structure, tracking_json_structure, existing_folder_option, ps)
    list_upload_files, main_total_generate_dataset_size = recursive_dataset_scan_for_ps(
        dataset_structure,
        tracking_json_structure,
        existing_file_option,
        list_upload_files,
        relative_path,
    )

    logger.info(f"Amount of files to upload: {len(list_upload_files)} ")


    # return and mark upload as completed if nothing is added to the manifest
    if len(list_upload_files) < 1:
        logger.warning("No files found to upload.")
        main_curate_progress_message = "No files were uploaded in this session"
        main_curate_status = "Done"
        return

    # 3. Add high-level metadata files to a list
    if "dataset_metadata" in soda.keys():
        existing_root_files = tracking_json_structure.get("children", {}).get("files", {})
        create_metadata_files_for_upload(
            soda, 
            list_upload_metadata_files,
            existing_root_files=existing_root_files,
            existing_file_option=existing_file_option,
            ps=ps
        )

    return {
        "list_upload_files": list_upload_files,
        "list_upload_metadata_files": list_upload_metadata_files,
        "list_upload_manifest_files": list_upload_manifest_files,
        "total_manifest_files": total_manifest_files,
        "main_total_generate_dataset_size": main_total_generate_dataset_size,
        "total_files": total_files,
        "total_metadata_files": total_metadata_files,
    }



def create_upload_manifest(soda, ps, ds):
    global logger

    # Progress tracking variables that are used for the frontend progress bar.
    global main_curate_progress_message
    global main_total_generate_dataset_size
    global main_generated_dataset_size
    global start_generate
    global main_initial_bfdataset_size
    global main_curation_uploaded_files
    global uploaded_folder_counter
    global current_size_of_uploaded_files
    global total_files
    global total_bytes_uploaded # current number of bytes uploaded to Pennsieve in the current session
    global client
    global files_uploaded
    global total_dataset_files
    global current_files_in_subscriber_session
    global renaming_files_flow
    global bytes_uploaded_per_file
    global total_bytes_uploaded_per_file
    global bytes_file_path_dict
    global elapsed_time
    global manifest_id
    global origin_manifest_id
    global main_curate_status
    global list_of_files_to_rename
    global renamed_files_counter




    total_files = 0
    total_dataset_files = 0
    total_metadata_files = 0
    total_manifest_files = 0
    main_curation_uploaded_files = 0
    total_bytes_uploaded = {"value": 0}
    total_bytes_uploaded_per_file = {}
    files_uploaded = 0
    renamed_files_counter = 0
    

    uploaded_folder_counter = 0
    current_size_of_uploaded_files = 0
    start = timer()
    try:
        def monitor_subscriber_progress(events_dict):
            """
            Monitors the progress of a subscriber and unsubscribes once the upload finishes. 
            """
            global files_uploaded
            global total_bytes_uploaded
            global bytes_uploaded_per_file
            global main_curation_uploaded_files
            global main_total_generate_dataset_size


            if events_dict["type"] == 1:  # upload status: file_id, total, current, worker_id
                file_id = events_dict["upload_status"].file_id
                total_bytes_to_upload = events_dict["upload_status"].total
                current_bytes_uploaded = events_dict["upload_status"].current

                status = events_dict["upload_status"].status
                if status == "2" or status == 2:
                    ps.unsubscribe(10)
                    logger.info("[UPLOAD COMPLETE EVENT RECEIVED]")
                    logger.info(f"Amount of bytes uploaded via sum: {sum(bytes_uploaded_per_file.values())} vs total bytes uploaded via difference: {total_bytes_uploaded['value']}")
                    logger.info(f"Amount of bytes Pennsieve Agent says via sum: {sum(bytes_uploaded_per_file.values())} vs amount of bytes we calculated before hand: {main_total_generate_dataset_size}")


                # only update the byte count if the current bytes uploaded is greater than the previous bytes uploaded
                # if current_bytes_uploaded > previous_bytes_uploaded:
                # update the file id's current total bytes uploaded value 
                bytes_uploaded_per_file[file_id] = current_bytes_uploaded
                total_bytes_uploaded["value"] = sum(bytes_uploaded_per_file.values())

                # check if the given file has finished uploading
                if current_bytes_uploaded == total_bytes_to_upload and  file_id != "":
                    files_uploaded += 1
                    main_curation_uploaded_files += 1



        # Set the Pennsieve Python Client's dataset to the Pennsieve dataset that will be uploaded to.
        selected_id = ds["content"]["id"]
        ps.use_dataset(selected_id)

        # Set variables needed throughout generation flow
        list_upload_files = []
        list_upload_metadata_files = []
        list_upload_manifest_files = []
        list_of_files_to_rename = {}
        brand_new_dataset = False
        dataset_structure = soda["dataset-structure"]
        generate_option = soda["generate-dataset"]["generate-option"]
        starting_point = soda["starting-point"]["origin"]
        relative_path = ds["content"]["name"]
 

        # 1. Scan the dataset structure and create a list of files/folders to be uploaded with the desired renaming
        if generate_option == "new" and starting_point == "new":
            logger.info("NO progress found so we will start from scratch and construct the manifest")
            main_curate_progress_message = "Preparing a list of files to upload"
            # we can assume no files/folders exist in the dataset since the generate option is new and starting point is also new
            # therefore, we can assume the dataset structure is the same as the tracking structure
            brand_new_dataset = True
            list_upload_files = create_upload_information_new(soda, relative_path)
                
            # For brand new datasets, no existing files to check - upload all metadata files
            logger.info("ps_upload_to_dataset: Creating metadata files for brand new dataset (no existing file checks needed)")
            create_metadata_files_for_upload(soda, list_upload_metadata_files)
        else:
            main_curate_progress_message = "Preparing a list of files to upload"
            info = create_upload_information_existing(soda, dataset_structure, ds, ps, relative_path)
            list_upload_files = info["list_upload_files"]
            list_upload_metadata_files = info["list_upload_metadata_files"]
            main_total_generate_dataset_size = info["main_total_generate_dataset_size"]
            total_files = info["total_files"]
            total_metadata_files = info["total_metadata_files"]
            brand_new_dataset = False

            


        # 2. Count how many files will be uploaded to inform frontend - do not count if we are resuming a previous upload that has made progress
        for folderInformation in list_upload_files:
            file_paths_count = len(folderInformation[0])
            total_files += file_paths_count
            total_dataset_files += file_paths_count


        # 3. Upload files and add to tracking list
        start_generate = 1

        
 
        if len(list_upload_files) <= 0:
            # TODO: Add information showing nothing added and no manifest created or maybe even just throw a 400 error
            logger.info("Manifest creation: Failed 0 files added to dataset.")
            raise EmptyDatasetError("The dataset you are trying to upload is empty.")
        
        
        main_curate_progress_message = ("Queuing dataset files for upload with the Pennsieve Agent..." + "<br>" + "This may take some time.")

        first_file_local_path = list_upload_files[0][0][0]

        if brand_new_dataset:
            first_relative_path = list_upload_files[0][4]
            first_final_name = list_upload_files[0][2][0]
        else:
            first_relative_path = list_upload_files[0][6]
            first_final_name = list_upload_files[0][4][0]

        # Extract folder_name (subfolder path) and high_lvl_folder from relative_path
        try:
            slash_idx = first_relative_path.index("/")
            folder_name = first_relative_path[slash_idx+1:]
            first_high_lvl_folder = first_relative_path[:slash_idx]
        except ValueError:
            # No slash in path - entire path is the high-level folder, no subfolders
            folder_name = ""
            first_high_lvl_folder = first_relative_path

        if first_final_name != basename(first_file_local_path):
            # if file name is not the same as local path, then it has been renamed in SODA
            if folder_name not in list_of_files_to_rename:
                list_of_files_to_rename[folder_name] = {"high_lvl_folder": first_high_lvl_folder}
            else:
                # Entry exists but may not have high_lvl_folder set
                if "high_lvl_folder" not in list_of_files_to_rename[folder_name]:
                    list_of_files_to_rename[folder_name]["high_lvl_folder"] = first_high_lvl_folder
            if basename(first_file_local_path) not in list_of_files_to_rename[folder_name]:
                list_of_files_to_rename[folder_name][basename(first_file_local_path)] = {
                    "final_file_name": first_final_name,
                    "id": "",
                }
                renamed_files_counter += 1

        manifest_data = ps.manifest.create(first_file_local_path, folder_name)
        manifest_id = manifest_data.manifest_id


        ums.set_df_mid(manifest_id)

        # remove the item just added to the manifest 
        list_upload_files[0][0].pop(0)

        # reset global variables used in the subscriber monitoring function
        bytes_uploaded_per_file = {}
        total_bytes_uploaded = {"value": 0}
        current_files_in_subscriber_session = total_dataset_files

        # there are files to add to the manifest if there are more than one file in the first folder or more than one folder
        if len(list_upload_files[0][0]) > 1 or len(list_upload_files) > 1:
            index_skip = True
            for folderInformation in list_upload_files:
                list_file_paths = folderInformation[0]
                if brand_new_dataset:
                    relative_path = folderInformation[4]
                    final_file_name_list = folderInformation[2]
                else:
                    relative_path = folderInformation[6]
                    final_file_name_list = folderInformation[4]
                # get the substring from the string relative_path that starts at the index of the / and contains the rest of the string
                try:
                    slash_idx = relative_path.index("/")
                    folder_name = relative_path[slash_idx+1:]
                    high_lvl_folder = relative_path[:slash_idx]
                except ValueError as e:
                    # No slash in path - entire path is the high-level folder, no subfolders
                    folder_name = ""
                    high_lvl_folder = relative_path

                # Add files to manfiest"
                final_files_index = 1 if index_skip else 0
                index_skip = False
                for file_path in list_file_paths:
                    file_file_name = final_file_name_list[final_files_index]
                    if file_file_name != basename(file_path):
                        # save the relative path, final name and local path of the file to be renamed
                        if folder_name not in list_of_files_to_rename:
                            list_of_files_to_rename[folder_name] = {"high_lvl_folder": high_lvl_folder}
                        else:
                            # Entry exists but may not have high_lvl_folder set
                            if "high_lvl_folder" not in list_of_files_to_rename[folder_name]:
                                list_of_files_to_rename[folder_name]["high_lvl_folder"] = high_lvl_folder
                        if basename(file_path) not in list_of_files_to_rename[folder_name]:
                            renamed_files_counter += 1
                            list_of_files_to_rename[folder_name][basename(file_path)] = {
                                "final_file_name": file_file_name,
                                "id": "",
                            }
                    ps.manifest.add(file_path, folder_name, manifest_id)
                    final_files_index += 1


        # add metadata files to the manifest
        if list_upload_metadata_files:
            current_files_in_subscriber_session += total_metadata_files
            # add the files to the manifest
            for manifest_path in list_upload_metadata_files:
                # subprocess call to the pennsieve agent to add the files to the manifest
                ps.manifest.add(manifest_path, target_base_path="", manifest_id=manifest_id)


        # add manifest files to the upload manifest
        if list_upload_manifest_files:
            current_files_in_subscriber_session += total_manifest_files
            for manifest_file_path in list_upload_manifest_files:
                # add the file to the manifest
                ps.manifest.add(manifest_file_path, "/", manifest_id)

        
        # set rename files to ums for upload resuming if this upload fails
        if renamed_files_counter > 0:
            ums.set_list_of_files_to_rename(list_of_files_to_rename)
            ums.set_rename_total_files(renamed_files_counter)


        # wait for all of the Agent's processes to finish to avoid errors when deleting files on Windows
        time.sleep(1)

        # at end of successful session reset tracking for folders created
        # main_curate_progress_message = "Success: COMPLETED!"
        # main_curate_status = "Done"
        main_curate_progress_message = "Success: MANIFEST CREATED!"
        main_curate_status = "MANIFEST STAGE DONE"

        
        # shutil.rmtree(manifest_folder_path) if isdir(manifest_folder_path) else 0
        end = timer()
        logger.info(f"Time for ps_upload_to_dataset function: {timedelta(seconds=end - start)}")

        return {
                "manifest_id": manifest_id, 
                "dataset_id": selected_id, 
                "list_of_files_to_rename": list_of_files_to_rename, 
                "size_of_dataset": main_total_generate_dataset_size, 
                "number_of_files": total_files
                }
    except Exception as e:
        logger.error(f"An error occurred in ps_upload_to_dataset function: {str(e)}")
        raise e



def rename_files_stage(ds):
    # 6. Rename files
    if list_of_files_to_rename:
        renaming_files_flow = True
        logger.info("ps_create_new_dataset (optional) step 8 rename files")
        logger.info("file-rename-fix-log: Entered rename step, list_of_files_to_rename keys: %s", list(list_of_files_to_rename.keys()))
        main_curate_progress_message = ("Preparing files to be renamed...")
        dataset_id = ds["content"]["id"]
        collection_ids = {}
        # gets the high level folders in the dataset
        r = requests.get(f"{PENNSIEVE_URL}/datasets/{dataset_id}", headers=create_request_headers(get_access_token()))
        logger.info("file-rename-fix-log: Requested dataset content for dataset_id: %s", dataset_id)
        r.raise_for_status()
        dataset_content = r.json()["children"]

        if dataset_content == []:
            logger.info("file-rename-fix-log: dataset_content is empty, entering wait loop")
            while dataset_content == []:
                logger.info("file-rename-fix-log: Waiting for dataset_content to be populated...")
                time.sleep(5)
                r = requests.get(f"{PENNSIEVE_URL}/datasets/{dataset_id}", headers=create_request_headers(get_access_token()))
                r.raise_for_status()
                dataset_content = r.json()["children"]

        collections_found = False
        logger.info("file-rename-fix-log: Starting search for high-level folders in dataset_content")
        collection_retry_count = 0
        max_collection_retries = 2  # 1 immediate retry after 5s, then proceed after 10s if still none
        while not collections_found and collection_retry_count < max_collection_retries:
            for item in dataset_content:
                if item["content"]["packageType"] == "Collection":
                    collections_found = True
                    collection_ids[item["content"]["name"]] = {"id": item["content"]["nodeId"]}
                    logger.info("file-rename-fix-log: Found collection: %s with id: %s", item['content']['name'], item['content']['nodeId'])

            if not collections_found:
                collection_retry_count += 1
                logger.info("file-rename-fix-log: No collections found, retrying after 10s... (attempt %d)", collection_retry_count)
                time.sleep(10)
                r = requests.get(f"{PENNSIEVE_URL}/datasets/{dataset_id}", headers=create_request_headers(get_access_token()))
                r.raise_for_status()
                dataset_content = r.json()["children"]

        if not collections_found:
            logger.info("file-rename-fix-log: Still no collections found after %d retries, proceeding with root-level file renaming.", max_collection_retries)

        for key in list_of_files_to_rename:
            # Key structure:
            # - key='' with high_lvl_folder set: Files directly in a high-level folder (e.g., primary/file.txt)
            # - key='' with high_lvl_folder='': Files at actual dataset root (no folder)
            # - key='subfolder' or 'sub1/sub2': Files in subfolders, high_lvl_folder derived from first part of key
            if key == '':
                # Files directly in high-level folder OR at dataset root - use stored high_lvl_folder
                high_lvl_folder_name = list_of_files_to_rename[key].get("high_lvl_folder", "")
                subfolder_amount = 0
                
                if high_lvl_folder_name and high_lvl_folder_name in collection_ids:
                    # Files directly in a high-level folder (e.g., primary/file.txt)
                    high_lvl_folder_id = collection_ids[high_lvl_folder_name]["id"]
                    list_of_files_to_rename[key]["id"] = high_lvl_folder_id
                    
                    # Get the high-level folder content
                    limit = 100
                    offset = 0
                    folder_content = []
                    while True:
                        r = requests.get(f"{PENNSIEVE_URL}/packages/{high_lvl_folder_id}?limit={limit}&offset={offset}", headers=create_request_headers(get_access_token()))
                        r.raise_for_status()
                        page = r.json().get("children", [])
                        folder_content.extend(page)
                        if len(page) < limit:
                            break
                        offset += limit
                    
                    # Find file IDs
                    for item in folder_content:
                        if item["content"]["packageType"] != "Collection":
                            file_name = item["content"]["name"]
                            file_id = item["content"]["nodeId"]
                            if file_name in list_of_files_to_rename[key]:
                                list_of_files_to_rename[key][file_name]["id"] = file_id
                else:
                    # Files at actual dataset root (no high-level folder)
                    list_of_files_to_rename[key]["id"] = dataset_id  # Use dataset_id for root-level lookup
                    
                    # Get dataset root content
                    limit = 100
                    offset = 0
                    root_content = []
                    while True:
                        r = requests.get(f"{PENNSIEVE_URL}/datasets/{dataset_id}?limit={limit}&offset={offset}", headers=create_request_headers(get_access_token()))
                        r.raise_for_status()
                        page = r.json().get("children", [])
                        root_content.extend(page)
                        if len(page) < limit:
                            break
                        offset += limit
                    
                    # Find file IDs at root
                    for item in root_content:
                        if item["content"]["packageType"] != "Collection":
                            file_name = item["content"]["name"]
                            file_id = item["content"]["nodeId"]
                            if file_name in list_of_files_to_rename[key]:
                                list_of_files_to_rename[key][file_name]["id"] = file_id
                continue
            
            # Non-empty key: derive high_lvl_folder from the key itself (first part before /)
            relative_path = key.split("/")
            high_lvl_folder_name = relative_path[0]
            subfolder_level = 0
            subfolder_amount = len(relative_path) - 1

            if high_lvl_folder_name in collection_ids:
                # subfolder_amount will be the amount of subfolders we need to call until we can get the file ID to rename

                high_lvl_folder_id = collection_ids[high_lvl_folder_name]["id"]
                limit = 100
                offset = 0
                dataset_content = []
                while True:
                    r = requests.get(f"{PENNSIEVE_URL}/packages/{high_lvl_folder_id}?limit={limit}&offset={offset}", headers=create_request_headers(get_access_token()))
                    r.raise_for_status()
                    page = r.json()["children"]
                    dataset_content.extend(page)

                    if len(page) < limit:
                        break
                    offset += limit

                if dataset_content == []:
                    # request until there is no children content, (folder is empty so files have not been processed yet)
                    while dataset_content == []:
                        time.sleep(3)
                        limit = 100 
                        offset = 0 

                        while True:
                            r = requests.get(f"{PENNSIEVE_URL}/packages/{high_lvl_folder_id}?limit={limit}&offset={offset}", headers=create_request_headers(get_access_token()))
                            r.raise_for_status()
                            page = r.json()["children"]
                            dataset_content.extend(page)
                            if len(page) < limit:
                                break
                            offset += limit
                        

                if subfolder_amount == 0:
                    # the file is in the high level folder
                    if "id" not in list_of_files_to_rename[key]:
                        # store the id of the folder to be used again in case the file id is not found (happens when not all files have been processed yet)
                        list_of_files_to_rename[key]["id"] = high_lvl_folder_id
                    
                    for item in dataset_content:
                        if item["content"]["packageType"] != "Collection":
                            file_name = item["content"]["name"]
                            file_id = item["content"]["nodeId"]

                            if file_name in list_of_files_to_rename[key]:
                                list_of_files_to_rename[key][file_name]["id"] = file_id
                else:
                    # file is within a subfolder and we recursively iterate until we get to the last subfolder needed
                    subfolder_id = collection_ids[high_lvl_folder_name]["id"]
                    while subfolder_level != subfolder_amount:
                        if dataset_content == []:
                            # subfolder has no content so request again
                            while dataset_content == []:
                                time.sleep(3)
                                limit = 100 
                                offset = 0
                                while True: 
                                    r = requests.get(f"{PENNSIEVE_URL}/packages/{subfolder_id}", headers=create_request_headers(get_access_token()))
                                    r.raise_for_status()
                                    page = r.json()["children"]
                                    dataset_content.extend(page)
                                    if len(page) < limit:
                                        break
                                    offset += limit
                    
                        for item in dataset_content:
                            if item["content"]["packageType"] == "Collection":
                                folder_name = item["content"]["name"]
                                folder_id = item["content"]["nodeId"]

                                if folder_name in relative_path:
                                    # we have found the folder we need to iterate through
                                    subfolder_level += 1

                                    limit = 100
                                    offset = 0 
                                    children = []
                                    while True: 
                                        r = requests.get(f"{PENNSIEVE_URL}/packages/{folder_id}?limit={limit}&offset={offset}", headers=create_request_headers(get_access_token()))
                                        r.raise_for_status()
                                        page = r.json()["children"]
                                        children.extend(page)
                                        if len(page) < limit:
                                            break
                                        offset += limit

                                    if subfolder_level != subfolder_amount:
                                        dataset_content = children
                                        if dataset_content == []:
                                            while dataset_content == []:
                                                # subfolder has no content so request again
                                                time.sleep(3)
                                                limit = 100
                                                offset = 0 
                                                while True: 
                                                    r = requests.get(f"{PENNSIEVE_URL}/packages/{folder_id}", headers=create_request_headers(get_access_token()))
                                                    r.raise_for_status()
                                                    page = r.json()["children"]
                                                    dataset_content.extend(page)
                                                    if len(page) < limit:
                                                        break
                                                    offset += limit
                                                
                                        subfolder_id = folder_id
                                        break
                                    else:
                                        # we are at the last folder in the relative path, we can get the file id
                                        if "id" not in list_of_files_to_rename[key]:
                                            # store the id of the last folder to directly call later in case not all files get an id
                                            list_of_files_to_rename[key]["id"] = folder_id
                                        for item in children:
                                            if item["content"]["packageType"] != "Collection":
                                                file_name = item["content"]["name"]
                                                file_id = item["content"]["nodeId"]

                                                if file_name in list_of_files_to_rename[key]:
                                                    # store the package id for renaming
                                                    list_of_files_to_rename[key][file_name]["id"] = file_id
                                else:
                                    continue

        # 8.5 Rename files - All or most ids have been fetched now rename the files or gather the ids again if not all files have been processed at this time
        main_curate_progress_message = "Renaming files..."
        main_generated_dataset_size = 0
        main_total_generate_dataset_size = renamed_files_counter
        for relative_path in list_of_files_to_rename:
            # Check if "id" exists for this path (may not exist if not yet set)
            if "id" not in list_of_files_to_rename[relative_path]:
                logger.warning(f"No 'id' key found for relative_path '{relative_path}', skipping...")
                continue
            
            collection_id = list_of_files_to_rename[relative_path]["id"]
            high_lvl_folder_name = list_of_files_to_rename[relative_path].get("high_lvl_folder", "")
            # Check if this is a dataset root file (key='' and no high_lvl_folder)
            is_dataset_root = (relative_path == '' and not high_lvl_folder_name)
            
            for file in list_of_files_to_rename[relative_path].keys():
                if file == "id" or file == "high_lvl_folder":
                    continue
                new_name = list_of_files_to_rename[relative_path][file]["final_file_name"]
                file_id = list_of_files_to_rename[relative_path][file]["id"]

                if file_id != "":
                    # id was found so make api call to rename with final file name
                    try:
                        r = requests.put(f"{PENNSIEVE_URL}/packages/{file_id}?updateStorage=true", json={"name": new_name}, headers=create_request_headers(get_access_token()))
                        r.raise_for_status()
                    except Exception as e:
                        if r.status_code == 500:
                            continue
                    main_generated_dataset_size += 1
                else:
                    # id was not found so keep trying to get the id until it is found
                    all_ids_found = False
                    retry_attempts = 0
                    max_retry_attempts = 60
                    while not all_ids_found and retry_attempts < max_retry_attempts:
                        retry_attempts += 1
                        time.sleep(3)

                        limit = 100
                        offset = 0
                        dataset_content = []

                        # Use correct endpoint: /datasets/ for root-level files, /packages/ for folder files
                        while True:
                            if is_dataset_root:
                                r = requests.get(f"{PENNSIEVE_URL}/datasets/{collection_id}?limit={limit}&offset={offset}", headers=create_request_headers(get_access_token()))
                            else:
                                r = requests.get(f"{PENNSIEVE_URL}/packages/{collection_id}?limit={limit}&offset={offset}", headers=create_request_headers(get_access_token()))
                            r.raise_for_status()
                            page = r.json().get("children", [])
                            dataset_content.extend(page)
                            if len(page) < limit:
                                break
                            offset += limit
                        
                        for item in dataset_content:
                            if item["content"]["packageType"] != "Collection":
                                file_name = item["content"]["name"]
                                found_file_id = item["content"]["nodeId"]

                                if file_name == file:
                                    # id was found so make api call to rename with final file name
                                    try:
                                        r = requests.put(f"{PENNSIEVE_URL}/packages/{found_file_id}?updateStorage=true", json={"name": new_name}, headers=create_request_headers(get_access_token()))
                                        r.raise_for_status()
                                    except Exception as e:
                                        if r.status_code == 500:
                                            continue
                                    main_generated_dataset_size += 1
                                    all_ids_found = True
                                    break
                    
                    if not all_ids_found:
                        logger.warning(f"Could not find file ID for '{file}' after {max_retry_attempts} attempts")


        


                # get the manifest id of the Pennsieve upload manifest created when uploading
  


main_curate_status = ""
main_curate_print_status = ""
main_curate_progress_message = ""
main_total_generate_dataset_size = 1
main_generated_dataset_size = 0
start_generate = 0
generate_start_time = 0
main_generate_destination = ""
main_initial_bfdataset_size = 0
myds = ""
renaming_files_flow = False
elapsed_time = None
manifest_id = None 
origin_manifest_id = None
curation_error_message = ""



def ps_check_dataset_files_validity(soda):
    """
    Minimal validation for the selected Pennsieve dataset.

    Args:
        soda: SODA request object with ps-dataset-selected

    Output:
        [] (no validation errors)
    """
    dataset_name = soda["ps-dataset-selected"]["dataset-name"]
    dataset_id = get_dataset_id(dataset_name)
    r = requests.get(f"{PENNSIEVE_URL}/datasets/{dataset_id}", headers=create_request_headers(get_access_token()))
    r.raise_for_status()
    root_folder = r.json()["children"]

    # if empty dataset, still considered valid in the context of stub validation
    if len(root_folder) == 0:
        return []

    return []
def check_server_access_to_files(file_list):
    # Return two lists, one that the server can open, and one that it can not.
    # This is to avoid the server trying to open files that it does not have access to.cf
    accessible_files = []
    inaccessible_files = []
    for file in file_list:
        if os.path.isfile(file) or os.path.isdir(file):
            accessible_files.append(file)
        else:
            inaccessible_files.append(file)

    return {"accessible_files": accessible_files, "inaccessible_files": inaccessible_files}


# TODO: Update for SDS 3.0
def clean_json_structure(soda):
    global logger
    # Delete any files on Pennsieve that have been marked as deleted
    def recursive_file_delete(folder):
        if "files" in folder.keys():
            for item in list(folder["files"]):
                if item in ["manifest.xlsx", "manifest.csv"]:
                    continue
                if "deleted" in folder["files"][item]["action"]:
                    # remove the file from the soda json structure
                    del folder["files"][item]

        for item in list(folder["folders"]):
            recursive_file_delete(folder["folders"][item])


    # Rename any files that exist on Pennsieve
    def recursive_file_rename(folder):
        if "files" in folder.keys():
            for item in list(folder["files"]):
                if (
                    "renamed" in folder["files"][item]["action"]
                    and folder["files"][item]["location"] == "ps"
                ):
                    continue

        for item in list(folder["folders"]):
            recursive_file_rename(folder["folders"][item])


    def recursive_folder_delete(folder):
        """
        Delete any stray folders that exist on Pennsieve
        Only top level files are deleted since the api deletes any
        files and folders that exist inside.
        """

        for folder_item in list(folder["folders"]):
            if folder["folders"][folder_item]["location"] == "ps":
                if "deleted" in folder["folders"][folder_item]["action"]:
                    del folder["folders"][folder_item]
                else:
                    recursive_folder_delete(folder["folders"][folder_item])
            else:
                recursive_folder_delete(folder["folders"][folder_item])
        return

    main_keys = soda.keys()
    dataset_structure = soda["dataset-structure"]

    if ("dataset-structure" not in main_keys and "dataset_metadata" not in main_keys):
        if "ps-dataset-selected" in main_keys:
            dataset_name = soda["ps-dataset-selected"]["dataset-name"]
        elif "generate-dataset" in main_keys:
            dataset_name = soda["generate-dataset"]["dataset-name"]
        else:
            dataset_name = "Unset Name"
        raise EmptyDatasetError(dataset_name)

    if "generate-dataset" in main_keys:
        # Check that local files/folders exist
        try:
            if error := check_local_dataset_files_validity(soda):
                raise LocalDatasetMissingSpecifiedFiles(error)
            # check that dataset is not empty after removing all the empty files and folders
            if not soda["dataset-structure"]["folders"] and "dataset_metadata" not in soda:
                if "ps-dataset-selected" in main_keys:
                    dataset_name = soda["ps-dataset-selected"]["dataset-name"]
                elif "generate-dataset" in main_keys:
                    dataset_name = soda["generate-dataset"]["dataset-name"]
                else:
                    dataset_name = "Unset Name"
                raise EmptyDatasetError(dataset_name)
        except Exception as e:
            raise e

    if "starting-point" in main_keys and soda["starting-point"][
        "origin"
    ] in ["ps", "local"]:
        recursive_file_delete(dataset_structure)
        recursive_folder_delete(dataset_structure)
        soda["dataset-structure"] = dataset_structure


    # here will be clean up the soda json object before creating the manifest file cards
    return {"soda": soda}



def validate_local_dataset_generate_path(soda):
    generate_dataset = soda["generate-dataset"]
    local_dataset_path = generate_dataset["path"]
    if not isdir(local_dataset_path):
        error_message = (
            "Error: The Path "
            + local_dataset_path
            + " is not found. Please select a valid destination folder for the new dataset"
        )
        raise FileNotFoundError(error_message)




def generating_on_ps(soda):
    return soda["generate-dataset"]["destination"] == "ps"

def uploading_with_ps_account(soda):
    return "ps-account-selected" in soda

def uploading_to_existing_ps_dataset(soda):
    return (soda.get("starting-point") is not None and soda["starting-point"].get("origin") == "ps")

def can_resume_prior_upload(resume_status):
    global ums 
    return resume_status and ums.df_mid_has_progress()

def generate_options_set(soda):
    return "generate-dataset" in soda.keys()


def get_dataset_with_backoff(selected_dataset_id):
    # check that dataset was created with a limited retry (for some users the dataset isn't automatically accessible)
    attempts = 0
    while(attempts < 3):
        try: 
            # whether we are generating a new dataset or merging, we want the dataset information for later steps
            r = requests.get(f"{PENNSIEVE_URL}/datasets/{selected_dataset_id}", headers=create_request_headers(get_access_token()))
            r.raise_for_status()
            return r.json()
        except Exception as e:
            attempts += 1 
            # check if final attempt
            if attempts >= 2:
                # raise the error to the user
                raise e
            time.sleep(10)


def generate_new_ds_ps_resume(soda, dataset_name, ps):
    # get the dataset id by the name 
    try: 
        selected_dataset_id = get_dataset_id(dataset_name)
    except Exception as e:
        if e.code == 404:
            # dataset does not exist - create it 
            ds = ps_create_new_dataset(dataset_name, ps)
            selected_dataset_id = ds["content"]["id"]
    
    myds = get_dataset_with_backoff(selected_dataset_id)
    ps_upload_to_dataset(soda, ps, myds, True)

def generate_new_ds_ps(soda, dataset_name, ps):
    ds = ps_create_new_dataset(dataset_name, ps)
    selected_dataset_id = ds["content"]["id"]    
    myds = get_dataset_with_backoff(selected_dataset_id)
    return myds


def manifest_file_factory(soda, ps):
    global main_generate_destination
    global main_total_generate_dataset_size

    main_generate_destination = soda["generate-dataset"]["destination"]

    logger.info("manifest_file_factory determining if generating manifest for new or existing dataset")
    logger.info(f"uploading_to_existing={uploading_to_existing_ps_dataset(soda)}")

    if uploading_to_existing_ps_dataset(soda):
        logger.info("PATH: Existing PS dataset")
        selected_dataset_id = get_dataset_id(
            soda["ps-dataset-selected"]["dataset-name"]
        )
        # make an api request to pennsieve to get the dataset details
        r = requests.get(f"{PENNSIEVE_URL}/datasets/{selected_dataset_id}", headers=create_request_headers(get_access_token()))
        r.raise_for_status()
        myds = r.json()

        clean_existing_ps_dataset(soda, myds)
        return create_upload_manifest(soda, ps, myds)
    else:
        logger.info("PATH: New PS Dataset")
        dataset_name = soda["generate-dataset"]["dataset-name"]

        logger.info("PATH: New PS Dataset -> checking if dataset exists")
        try: 
            selected_dataset_id = get_dataset_id(dataset_name)
            logger.info(f"PATH: New PS Dataset ->  dataset exists ({selected_dataset_id}) -> ps_upload_to_dataset")
        except Exception as e:
            if isinstance(e, PennsieveDatasetCannotBeFound):
                logger.info("PATH: New PS Dataset ->  dataset not found -> create_upload_manifest")
                myds =  generate_new_ds_ps(soda, dataset_name, ps)
                logger.info("New dataset flow - initiating manifest creation")
                return create_upload_manifest(soda, ps, myds)
            else:
                raise Exception(f"{e.status_code}, {e.message}")
        myds = get_dataset_with_backoff(selected_dataset_id)
    
        return create_upload_manifest(soda, ps, myds)

                        
                


def validate_dataset_structure(soda):

    global main_curate_status
    global main_curate_progress_message
    global logger

    # 1] Check for potential errors
    logger.info("main_curate_function step 1")

    if not generate_options_set(soda):
        main_curate_status = "Done"
        raise GenerateOptionsNotSet()

    # 1.1. If the dataset is being generated locally then check that the local destination is valid
    if generating_locally(soda): 
        main_curate_progress_message = "Checking that the local destination selected for generating your dataset is valid"
        try: 
            validate_local_dataset_generate_path(soda)
        except Exception as e:
            main_curate_status = "Done"
            raise e
        

    logger.info("main_curate_function step 1.2")

    # 1.2. If generating dataset to Pennsieve or any other Pennsieve actions are requested check that the destination is valid
    if uploading_with_ps_account(soda):
        # check that the Pennsieve account is valid
        try: 
            main_curate_progress_message = (
                "Checking that the selected Pennsieve account is valid"
            )
            accountname = soda["ps-account-selected"]["account-name"]
            connect_pennsieve_client(accountname)
        except Exception as e:
            main_curate_status = "Done"
            raise e

    if uploading_to_existing_ps_dataset(soda):
        # check that the Pennsieve dataset is valid
        try:
            main_curate_progress_message = (
                "Checking that the selected Pennsieve dataset is valid"
            )
            bfdataset = soda["ps-dataset-selected"]["dataset-name"]
            selected_dataset_id = get_dataset_id(bfdataset)

        except Exception as e:
            main_curate_status = "Done"
            bfdataset = soda["ps-dataset-selected"]["dataset-name"]
            raise PennsieveDatasetCannotBeFound(bfdataset)

        # check that the user has permissions for uploading and modifying the dataset
        main_curate_progress_message = "Checking that you have required permissions for modifying the selected dataset"
        role = pennsieve_get_current_user_permissions(selected_dataset_id, get_access_token())["role"]
        if role not in ["owner", "manager", "editor"]:
            main_curate_status = "Done"
            raise PennsieveActionNoPermission("uploading to Pennsieve dataset")

    logger.info("main_curate_function step 1.3")


    # 1.3. Check that specified dataset files and folders are valid (existing path) if generate dataset is requested
    # Note: Empty folders and 0 kb files will be removed without warning (a warning will be provided on the front end before starting the curate process)
    # Check at least one file or folder are added to the dataset
    main_curate_progress_message = "Checking that the dataset is not empty"
    if virtual_dataset_empty(soda):
        main_curate_status = "Done" 
        if "generate-options" in soda.keys():
            dataset_name = soda["generate-options"]["dataset-name"]
        elif "ps-dataset-selected" in soda.keys():
            dataset_name = soda["ps-dataset-selected"]["dataset-name"]
        else:
            dataset_name = "Name not set"
        raise EmptyDatasetError(dataset_name)


    logger.info("main_curate_function step 1.3.1")

    # Check that local files/folders exist
    if error := check_local_dataset_files_validity(soda):
        main_curate_status = "Done"
        raise LocalDatasetMissingSpecifiedFiles(error)


    # check that dataset is not empty after removing all the empty files and folders
    if virtual_dataset_empty(soda):
        main_curate_status = "Done"
        if "generate-options" in soda.keys():
            dataset_name = soda["generate-options"]["dataset-name"]
        elif "ps-dataset-selected" in soda.keys():
            dataset_name = soda["ps-dataset-selected"]["dataset-name"]
        else:
            dataset_name = "Name not set"
        raise EmptyDatasetError(dataset_name, "The dataset is empty after removing all the empty files and folders.")


    logger.info("main_curate_function step 1.3.2")
    # Check that bf files/folders exist (Only used for when generating from an existing Pennsieve dataset)
    if uploading_to_existing_ps_dataset(soda):                     
        try:
            main_curate_progress_message = (
                "Checking that the Pennsieve files and folders are valid"
            )
            if soda["generate-dataset"]["destination"] == "ps":
                if error := ps_check_dataset_files_validity(soda):
                    logger.info("Failed to validate dataset files")
                    logger.info(error)
                    main_curate_status = "Done"
                    raise PennsieveDatasetFilesInvalid(error)
        except Exception as e:
            main_curate_status = "Done"
            raise e



def reset_upload_session_environment(resume):
    global main_curate_status
    global main_curate_progress_message
    global main_total_generate_dataset_size
    global main_generated_dataset_size
    global start_generate
    global generate_start_time
    global main_generate_destination
    global main_initial_bfdataset_size
    global main_curation_uploaded_files
    global uploaded_folder_counter
    global ums

    global myds
    global generated_dataset_id
    global bytes_file_path_dict
    global renaming_files_flow
    global curation_error_message

    start_generate = 0
    myds = ""

    generate_start_time = time.time()

    # variables for tracking the progress of the curate process on the frontend 
    main_curate_status = ""
    main_curate_progress_message = "Starting..."
    main_total_generate_dataset_size = 0
    main_generated_dataset_size = 0
    main_curation_uploaded_files = 0
    uploaded_folder_counter = 0
    generated_dataset_id = None
    curation_error_message = ""

    main_curate_status = "Curating"
    main_curate_progress_message = "Starting dataset curation"
    main_generate_destination = ""
    main_initial_bfdataset_size = 0

    if not resume:
        ums.set_df_mid(None)
        ums.set_elapsed_time(None)
        ums.set_total_files_to_upload(0)
        ums.set_main_total_generate_dataset_size(0)
        # reset the rename information back to default
        ums.set_renaming_files_flow(False) # this determines if we failed while renaming files after the upload is complete
        ums.set_rename_total_files(None)
        ums.set_list_of_files_to_rename(None)
        renaming_files_flow = False
        # reset the calculated values for the upload session
        bytes_file_path_dict = {}




def create_upload_manifest_pipeline(soda):
    global logger
    global main_curate_status
    global manifest_id 
    global origin_manifest_id
    global total_files
    global curation_error_message
    global main_curation
    global main
    global main_curate_progress_message
    global ps

    logger.info("Creating upload manifest")
    logger.info(f"Generating dataset metadata generate-options={soda['generate-dataset']}")

    validate_dataset_structure(soda)

    logger.info("Generating dataset step 3")


    # 2] Generate
    main_curate_progress_message = "Preparing dataset for upload..."
    try:
        logger.info("Creating Pennsieve manifest")
        accountname = soda["ps-account-selected"]["account-name"]
        ps = connect_pennsieve_client(accountname)
        return manifest_file_factory(soda, ps)
    except Exception as e:
        logger.error("Error creating manifest file")
        main_curate_status = "Done"
        curation_error_message = str(e)
        raise e



def main_curate_function_progress():
    """
    Function frequently called by front end to help keep track of the dataset generation progress
    """

    global main_curate_status  # empty if curate on going, "Done" when main curate function stopped (error or completed)
    global main_curate_progress_message
    global main_total_generate_dataset_size
    global main_generated_dataset_size
    global start_generate
    global generate_start_time
    global main_generate_destination
    global main_initial_bfdataset_size
    global main_curation_uploaded_files
    global total_bytes_uploaded # current number of bytes uploaded to Pennsieve in the upload session
    global myds
    global renaming_files_flow
    global ums 
    global elapsed_time
    global curation_error_message


    prior_elapsed_time = ums.get_elapsed_time()
    if prior_elapsed_time is not None: 
        elapsed_time =  ( time.time() - generate_start_time ) + prior_elapsed_time
    else:
        elapsed_time = time.time() - generate_start_time

    elapsed_time_formatted = time_format(elapsed_time)


    if renaming_files_flow:
        testing_variable = main_generated_dataset_size
    else:
        testing_variable = total_bytes_uploaded["value"]

    return {
        "main_curate_status": main_curate_status,
        "start_generate": start_generate,
        "main_curate_progress_message": main_curate_progress_message,
        "main_total_generate_dataset_size": main_total_generate_dataset_size,
        "main_generated_dataset_size": testing_variable,
        "elapsed_time_formatted": elapsed_time_formatted,
        "total_files_uploaded": main_curation_uploaded_files,
        "generated_dataset_id": myds["content"]["id"] if myds != "" else None, # when a new dataset gets generated log its id to our analytics
        "generated_dataset_int_id": myds["content"]["intId"] if myds != "" else None,
        "curation_error_message": curation_error_message,
    }


def preview_dataset(soda):
    """
    Associated with 'Preview' button in the SODA interface
    Creates a folder for preview and adds mock files based on the files specified in the UI by the user (same name as origin but 0 kb in size)
    Opens the dialog box to showcase the files / folders added

    Args:
        soda: soda dict with information about all specified files and folders
    Action:
        Opens the dialog box at preview_path
    Returns:
        preview_path: path of the folder where the preview files are located
    """

    preview_path = join(userpath, "SODA", "Preview_dataset")

    # remove empty files and folders from dataset
    try:
        check_empty_files_folders(soda)
    except Exception as e:
        raise e

    # create Preview_dataset folder
    try:
        if isdir(preview_path):
            shutil.rmtree(preview_path, ignore_errors=True)
        makedirs(preview_path)
    except Exception as e:
        raise e

    try:

        if "dataset-structure" in soda.keys():
            # create folder structure
            def recursive_create_mock_folder_structure(my_folder, my_folderpath):
                if "folders" in my_folder.keys():
                    for folder_key, folder in my_folder["folders"].items():
                        folderpath = join(my_folderpath, folder_key)
                        if not isdir(folderpath):
                            mkdir(folderpath)
                        recursive_create_mock_folder_structure(folder, folderpath)

                if "files" in my_folder.keys():
                    for file_key, file in my_folder["files"].items():
                        if "deleted" not in file["action"]:
                            open(join(my_folderpath, file_key), "a").close()

            dataset_structure = soda["dataset-structure"]
            folderpath = preview_path
            recursive_create_mock_folder_structure(dataset_structure, folderpath)

            if "manifest-files" in soda.keys() and "folders" in dataset_structure.keys():
                for folder_key, folder in dataset_structure["folders"].items():
                    manifest_path = join(preview_path, folder_key, "manifest.xlsx")
                    if not isfile(manifest_path):
                        open(manifest_path, "a").close()

        if "metadata-files" in soda.keys():
            for metadata_key in soda["metadata-files"].keys():
                open(join(preview_path, metadata_key), "a").close()

        if len(listdir(preview_path)) > 0:
            folder_in_preview = listdir(preview_path)[0]
            open_file(join(preview_path, folder_in_preview))
        else:
            open_file(preview_path)

        return preview_path

    except Exception as e:
        raise e


def generate_manifest_file_locally(generate_purpose, soda):
    """
    Function to generate manifest files locally
    """


    global manifest_folder_path

    def recursive_item_path_create(folder, path):
        if "files" in folder.keys():
            for item in list(folder["files"]):
                if "folderpath" not in folder["files"][item]:
                    folder["files"][item]["folderpath"] = path[:]

        if "folders" in folder.keys():
            for item in list(folder["folders"]):
                if "folderpath" not in folder["folders"][item]:
                    folder["folders"][item]["folderpath"] = path[:]
                    folder["folders"][item]["folderpath"].append(item)
                recursive_item_path_create(
                    folder["folders"][item], folder["folders"][item]["folderpath"][:]
                )

        return

    def copytree(src, dst, symlinks=False, ignore=None):
        for item in os.listdir(src):
            s = os.path.join(src, item)
            d = os.path.join(dst, item)
            if os.path.isdir(s):
                if os.path.exists(d):
                    shutil.rmtree(d)
                shutil.copytree(s, d, symlinks, ignore)
            else:
                shutil.copy2(s, d)

    dataset_structure = soda["dataset-structure"]
    manifest_destination = soda["manifest-files"]["local-destination"]

    recursive_item_path_create(dataset_structure, [])
    create_high_lvl_manifest_files_existing_ps_starting_point(soda, manifest_folder_path)

    if generate_purpose == "edit-manifest":
        manifest_destination = os.path.join(manifest_destination, "manifest_file")

    else:
        manifest_destination = return_new_path(
            os.path.join(manifest_destination, "manifest_file")
        )

        copytree(manifest_folder_path, manifest_destination)



    if generate_purpose == "edit-manifest":
        return {"success_message_or_manifest_destination": manifest_destination}

    open_file(manifest_destination)
    return {"success_message_or_manifest_destination": "success"}



def generate_manifest_file_data(dataset_structure):
    # Helper: Determine file extension (handles double extensions).
    # ``ps_recognized_file_extensions`` may contain extensions where one is a
    # suffix of another (for example ".tar" and ".tar.gz").  If we iterate
    # naively we can return the shorter extension and strip the wrong suffix.
    # To avoid this we sort the list by length (longest first) and perform a
    # case-insensitive comparison.
    def get_file_extension(filename):
        lower_name = filename.lower()
        for ext in sorted(ps_recognized_file_extensions, key=len, reverse=True):
            if lower_name.endswith(ext.lower()):
                return ext
        # fallback to the standard splitext if nothing matches
        return os.path.splitext(filename)[1]

    # Helper: Create a manifest row for a folder
    def create_folder_entry(folder_name, path_parts):
        full_path = "/".join(path_parts + [folder_name]) + "/"
        return [
            full_path.lstrip("/"),
            "", "", "folder", "", "", "", "", "", ""
        ]

    # Helper: Create a manifest row for a file
    def create_file_entry(file_name, file_info, path_parts, timestamp):
        entry = [
            "/".join(path_parts + [file_name]).lstrip("/"),
            timestamp,
            file_info["description"],
            get_file_extension(file_name),
            "", "", "", "", "",
            file_info.get("additional-metadata", "")
        ]

        # Append any extra columns dynamically
        if "extra_columns" in file_info:
            for key, value in file_info["extra_columns"].items():
                entry.append(value)
                if key not in header_row:
                    header_row.append(key)

        return entry

    # Recursive traversal of folders and files
    def traverse_folders(folder, path_parts):
        if not manifest_data:
            manifest_data.append(header_row)

        # Process files
        for file_name, file_info in folder.get("files", {}).items():
            file_path = file_info.get("path")
            if not file_path:
                continue
            if file_name in {"manifest.xlsx", "manifest.csv"}:
                continue

            if file_info["location"] == "ps":
                timestamp = file_info["timestamp"]
            else:
                local_path = pathlib.Path(file_info["path"])
                # Create proper ISO 8601 timestamp
                dt = datetime.fromtimestamp(local_path.stat().st_mtime, tz=timezone.utc)
                # per the SDS spec, replace '.' with ',' in the timestamp fractional seconds section
                timestamp = dt.isoformat().replace(".", ",").replace("+00:00", "Z")


            manifest_data.append(create_file_entry(file_name, file_info, path_parts, timestamp))

        # Process subfolders
        for subfolder_name, subfolder in folder.get("folders", {}).items():
            manifest_data.append(create_folder_entry(subfolder_name, path_parts))
            traverse_folders(subfolder, path_parts + [subfolder_name])

    # Initialize manifest data and header
    manifest_data = []
    header_row = [
        "filename", "timestamp", "description", "file type", "entity",
        "data modality", "also in dataset", "data dictionary path",
        "entity is transitive", "Additional Metadata"
    ]
    local_timezone = TZLOCAL()

    traverse_folders(dataset_structure, [])

    return manifest_data




